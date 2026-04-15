"""
Shared fixtures for Planificador de Talleres test suite.

Uses the REAL database (Neon PostgreSQL) with test-safe data.
Test trimestre: "TEST-Q1" - cleaned up after each test.
"""

import pytest
import pytest_asyncio
from httpx import AsyncClient, ASGITransport
from sqlalchemy import text

# Import the FastAPI app
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from main import app
from app.db import AsyncSessionLocal

# ── Constants ────────────────────────────────────────────────

TEST_TRIMESTRE = "TEST-Q1"
TEST_EMPRESA_PREFIX = "TEST_EMPRESA_"


# ── Fixtures ─────────────────────────────────────────────────

@pytest_asyncio.fixture
async def client():
    """Async HTTP client for testing."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac


@pytest_asyncio.fixture
async def db_session():
    """Direct database session for test setup/cleanup."""
    async with AsyncSessionLocal() as session:
        yield session


@pytest_asyncio.fixture(autouse=True)
async def cleanup(db_session):
    """
    Clean up test data after each test.
    Runs automatically after every test.
    """
    yield  # Test runs here

    # Cleanup all test data
    await _cleanup_test_data(db_session)


async def _cleanup_test_data(db):
    """Remove all test trimestre data from the database."""
    # Order matters due to foreign keys
    cleanup_queries = [
        # Delete from tables that reference planificacion first
        f'DELETE FROM planificacion WHERE trimestre = \'{TEST_TRIMESTRE}\'',
        f'DELETE FROM frecuencia WHERE trimestre = \'{TEST_TRIMESTRE}\'',
        f'DELETE FROM "configTrimestral" WHERE trimestre = \'{TEST_TRIMESTRE}\'',
        f'DELETE FROM "historicoTaller" WHERE trimestre = \'{TEST_TRIMESTRE}\'',
        f'DELETE FROM "solverLog" WHERE trimestre = \'{TEST_TRIMESTRE}\'',
        # Delete test empresas (if any were created)
        f'DELETE FROM empresa WHERE nombre LIKE \'{TEST_EMPRESA_PREFIX}%\'',
        # Delete test festivos
        f'DELETE FROM festivo WHERE trimestre = \'{TEST_TRIMESTRE}\'',
    ]

    for query in cleanup_queries:
        try:
            await db.execute(text(query))
        except Exception:
            pass  # Table might not exist or no rows to delete

    await db.commit()


# ── Helper Functions (available to all tests) ────────────────

def create_test_excel(slots_data: list[dict]) -> bytes:
    """
    Creates an Excel file in the expected calendar format for import testing.

    Args:
        slots_data: List of dicts with slot info (semana, dia, taller, empresa, estado, etc.)

    Returns:
        Excel file as bytes
    """
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calendario"

    headers = [
        "Semana", "Fecha", "Dia", "Horario", "Turno", "Empresa",
        "Taller", "Programa", "Ciudad", "Tipo", "Estado", "Confirmado"
    ]
    ws.append(headers)

    for slot in slots_data:
        ws.append([
            slot.get("semana", 1),
            slot.get("fecha", ""),
            slot.get("dia", "L"),
            slot.get("horario", "09:00-10:30"),
            slot.get("turno", "M"),
            slot.get("empresa", ""),
            slot.get("taller", "EF Lunes M 1"),
            slot.get("programa", "EF"),
            slot.get("ciudad", "MADRID"),
            slot.get("tipo", "BASE"),
            slot.get("estado", "PLANIFICADO"),
            slot.get("confirmado", ""),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def create_festivos_excel(festivos: list[dict]) -> bytes:
    """
    Creates an Excel file with festivos data for import testing.

    Args:
        festivos: List of dicts with fecha and motivo

    Returns:
        Excel file as bytes
    """
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()

    # Empresas sheet (minimal, required for import endpoint)
    ws_emp = wb.active
    ws_emp.title = "Empresas"
    ws_emp.append(["nombre", "tipo"])

    # Festivos sheet
    ws_fest = wb.create_sheet("Festivos")
    ws_fest.append(["Fecha", "Motivo"])

    for f in festivos:
        ws_fest.append([f.get("fecha"), f.get("motivo", "")])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ── Test Data Helpers ────────────────────────────────────────

async def setup_test_config_trimestral(db, trimestre: str = TEST_TRIMESTRE, empresa_ids: list[int] = None):
    """
    Creates configTrimestral entries for test trimestre.
    If empresa_ids not provided, uses first 5 active empresas.
    """
    if empresa_ids is None:
        # Get first 5 active empresas
        result = await db.execute(
            text("SELECT id FROM empresa WHERE activa = true ORDER BY id LIMIT 5")
        )
        empresa_ids = [row["id"] for row in result.mappings().all()]

    for eid in empresa_ids:
        await db.execute(
            text("""
                INSERT INTO "configTrimestral" (
                    "empresaId", trimestre, "tipoParticipacion",
                    "disponibilidadDias", "updatedAt"
                )
                VALUES (:eid, :tri, 'AMBAS', 'L,M,X,J,V', NOW())
                ON CONFLICT ("empresaId", trimestre) DO NOTHING
            """),
            {"eid": eid, "tri": trimestre}
        )

    await db.commit()
    return empresa_ids


async def setup_test_frecuencias(db, trimestre: str = TEST_TRIMESTRE, empresa_ids: list[int] = None):
    """
    Creates confirmed frecuencias for test trimestre.
    Each empresa gets 2 EF + 1 IT for simplicity.
    """
    if empresa_ids is None:
        # Get empresas from configTrimestral
        result = await db.execute(
            text("""
                SELECT "empresaId" FROM "configTrimestral"
                WHERE trimestre = :tri
            """),
            {"tri": trimestre}
        )
        empresa_ids = [row["empresaId"] for row in result.mappings().all()]

    for eid in empresa_ids:
        # Get config id
        cfg_result = await db.execute(
            text("""
                SELECT id FROM "configTrimestral"
                WHERE "empresaId" = :eid AND trimestre = :tri
            """),
            {"eid": eid, "tri": trimestre}
        )
        cfg = cfg_result.mappings().first()
        if not cfg:
            continue

        await db.execute(
            text("""
                INSERT INTO frecuencia (
                    "configId", "empresaId", trimestre,
                    "talleresEF", "talleresIT", "totalAsignado",
                    "semaforoCalculado", "scoreCalculado", "esNueva"
                )
                VALUES (
                    :cfg_id, :eid, :tri,
                    2, 1, 3,
                    'VERDE', 80.0, false
                )
                ON CONFLICT ("empresaId", trimestre) DO NOTHING
            """),
            {"cfg_id": cfg["id"], "eid": eid, "tri": trimestre}
        )

    await db.commit()
