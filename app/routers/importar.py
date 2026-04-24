"""
IMPORTACIÓN MASIVA — Carga de datos desde Excel maestro (v2)
FIX: comparación case-insensitive para evitar duplicados.

Endpoints:
  1. POST /api/importar/empresas   → upsert empresa + empresaCiudad + configTrimestral
  2. POST /api/importar/historico   → importar histórico desde Excel Q1
  3. GET  /api/importar/estado/{t}  → resumen de datos cargados

Tablas Prisma (camelCase → SQL con comillas dobles).
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from pydantic import BaseModel
from io import BytesIO

from app.db import get_db

router = APIRouter()


# ── Schemas ──────────────────────────────────────────────────


class ImportEmpresasResult(BaseModel):
    total_empresas: int
    creadas: int
    actualizadas: int
    ciudades_creadas: list[str]
    empresa_ciudad_links: int
    config_trimestral_creadas: int
    configs_actualizadas: int  # NEW: how many configs were updated vs created
    trimestre_target: str      # NEW: which trimestre was targeted
    festivos_importados: int
    warnings: list[str]


class ImportHistoricoResult(BaseModel):
    trimestre: str
    total_filas: int
    importadas_ok: int
    importadas_cancelado: int
    vacantes_ignoradas: int
    empresas_no_encontradas: list[str]
    talleres_no_encontrados: list[str]
    warnings: list[str]
    
 


# ── Constantes ───────────────────────────────────────────────

CIUDAD_ABREV = {
    "MAD": "MADRID",
    "MADRID": "MADRID",
    "BCN": "BARCELONA",
    "BARCELONA": "BARCELONA",
    "VLC": "VALENCIA",
    "VALENCIA": "VALENCIA",
    "ZGZ": "ZARAGOZA",
    "ZARAGOZA": "ZARAGOZA",
    "SEV": "SEVILLA",
    "SEVILLA": "SEVILLA",
    "MAL": "MALAGA",
    "MÁLAGA": "MALAGA",
    "MALAGA": "MALAGA",
    "MALL": "MALLORCA",
    "MALLORCA": "MALLORCA",
    "CAN": "CANARIAS",
    "CANARIAS": "CANARIAS",
}

CIUDAD_COLUMNS = {"mad", "bcn", "vlc", "zgz", "sev", "mal", "málag", "mall", "can"}

HEADER_MAP = {
    "nombre": "nombre",
    "tipo": "tipo",
    "semaforo": "semaforo",
    "semáforo": "semaforo",
    "scorev3": "scoreV3",
    "score": "scoreV3",
    "score_v3": "scoreV3",
    "fiabilidadreciente": "fiabilidadReciente",
    "fiabilidad": "fiabilidadReciente",
    "escomodin": "esComodin",
    "comodin": "esComodin",
    "aceptaextras": "aceptaExtras",
    "extras": "aceptaExtras",
    "maxextrastrimestre": "maxExtrasTrimestre",
    "max_extras": "maxExtrasTrimestre",
    "prioridadreduccion": "prioridadReduccion",
    "prioridad": "prioridadReduccion",
    "tienebolsa": "tieneBolsa",
    "bolsa": "tieneBolsa",
    "turnopreferido": "turnoPreferido",
    "turno": "turnoPreferido",
    "activa": "activa",
    "esnueva": "esNueva",
    "nueva": "esNueva",
    "notas": "notas",
    # Frecuencia solicitada — OLD format (single column = total)
    "frecuenciasolicitada": "frecuenciaSolicitada",
    "frecuencia": "frecuenciaSolicitada",
    "freq": "frecuenciaSolicitada",
    "talleres": "frecuenciaSolicitada",
    # Frecuencia EF/IT — NEW format (split columns)
    "frecuenciaef": "frecuenciaEF",
    "freq_ef": "frecuenciaEF",
    "tallersef": "frecuenciaEF",
    "ef": "frecuenciaEF",
    "frecuenciait": "frecuenciaIT",
    "freq_it": "frecuenciaIT",
    "talleresit": "frecuenciaIT",
    "it": "frecuenciaIT",
    # escuelaPropia (V18)
    "escuelapropia": "escuelaPropia",
    "escuela_propia": "escuelaPropia",
    "escuela": "escuelaPropia",
    # disponibilidadDias (V18)
    "disponibilidaddias": "disponibilidadDias",
    "disponibilidad": "disponibilidadDias",
    "dias": "disponibilidadDias",
    "dias_disponibles": "disponibilidadDias",
    # voluntariosDisponibles (V18)
    "voluntariosdisponibles": "voluntariosDisponibles",
    "voluntarios": "voluntariosDisponibles",
}

# Normalización empresas Q1
NORMALIZE_EMPRESA_Q1 = {
    "ACCIONA ": "ACCIONA",
    "AECOM ": "AECOM",
    "ATREVIA ": "ATREVIA",
    "BANKINTER": "BANKINTER",
    " INDRA": "INDRA",
    "MP CAPGEMINI": "CAPGEMINI",
    "MP ENDESA": "ENDESA",
    "MP LDA": "LDA",
    "MP AON": "F. AON",
    "MP SANTANDER": "B.SANTANDER",
    "SERVITEC / CAPGEMINI": "SERVITEC",
    "SERVITEC / SERVITEC": "SERVITEC",
    "URBASER / SACYR": "URBASER",
}


# ── Parsers ──────────────────────────────────────────────────


def _bool(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("SI", "SÍ", "YES", "TRUE", "1", "X")


def _float(val, d=0.0) -> float:
    if val is None or str(val).strip() == "":
        return d
    try:
        return float(val)
    except:  # noqa: E722
        return d


def _int(val, d=0) -> int:
    if val is None or str(val).strip() == "":
        return d
    try:
        return int(float(val))
    except:  # noqa: E722
        return d


def _str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _semaforo(val) -> str:
    if val is None:
        return "AMBAR"
    s = str(val).strip().upper().replace("Á", "A")
    return s if s in ("VERDE", "AMBAR", "ROJO") else "AMBAR"


def _programa(val) -> str:
    if val is None:
        return "AMBAS"
    s = str(val).strip().upper()
    return s if s in ("EF", "IT", "AMBAS") else "AMBAS"


def _prioridad(val) -> str:
    if val is None:
        return "MEDIA"
    s = str(val).strip().upper()
    return s if s in ("ALTA", "MEDIA", "BAJA") else "MEDIA"


_DIAS_CANONICOS = ["L", "M", "X", "J", "V"]


def _normalizar_dias(val) -> str | None:
    """Coerces day input to canonical 'L,M,X,J,V'-style. Returns None if empty
    or if any non-day character remains after stripping separators (strict)."""
    if val is None:
        return None
    if isinstance(val, (list, tuple, set)):
        items = [str(x) for x in val]
    else:
        items = [str(val)]
    chars: list[str] = []
    for item in items:
        s = item.upper()
        for sep in (",", ";", "-", " ", "/", "."):
            s = s.replace(sep, "")
        chars.extend(s)
    if not chars:
        return None
    if not all(c in _DIAS_CANONICOS for c in chars):
        return None
    return ",".join(d for d in _DIAS_CANONICOS if d in set(chars))


def _to_int_or_none(val) -> int | None:
    """Coerces to int; returns None on missing/blank/unparseable."""
    if val is None:
        return None
    if isinstance(val, str) and val.strip() == "":
        return None
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


# ══════════════════════════════════════════════════════════════
# ENDPOINT 1: Importar empresas
# ══════════════════════════════════════════════════════════════


@router.post("/empresas", response_model=ImportEmpresasResult)
async def importar_empresas(
    file: UploadFile = File(...),
    trimestre: str = Form("2026-Q2"),
    db: AsyncSession = Depends(get_db),
):
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    warnings: list[str] = []

    # ── Buscar hoja Empresas ─────────────────────────────────
    sheet_map = {s.lower(): s for s in wb.sheetnames}
    ws_name = sheet_map.get("empresas") or sheet_map.get("empresa")
    if not ws_name:
        raise HTTPException(
            400, f"Hoja 'Empresas' no encontrada. Hojas: {wb.sheetnames}"
        )
    ws = wb[ws_name]

    # ── Detectar columnas ────────────────────────────────────
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    col_map: dict[str, int] = {}
    ciudad_cols: dict[str, int] = {}
    ciudades_csv_col: int | None = None

    for idx, h in enumerate(headers):
        h_clean = h.replace(" ", "").replace("_", "").lower()
        if h_clean in HEADER_MAP:
            col_map[HEADER_MAP[h_clean]] = idx
        elif h_clean in CIUDAD_COLUMNS or h.upper() in CIUDAD_ABREV:
            ciudad_cols[h.upper()] = idx
        elif h_clean == "ciudades":
            ciudades_csv_col = idx

    if "nombre" not in col_map:
        raise HTTPException(400, f"Columna 'nombre' no encontrada. Headers: {headers}")

    # Warn about score columns being ignored (auto-calculated now)
    score_cols_found = [c for c in ["scoreV3", "semaforo", "fiabilidadReciente"] if c in col_map]
    if score_cols_found:
        warnings.append(
            f"Columnas {score_cols_found} ignoradas — los scores se calculan automáticamente desde el histórico"
        )

    # ── Leer filas ───────────────────────────────────────────
    empresas_data: list[dict] = []
    seen_names: set[str] = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        nombre_raw = (
            row[col_map["nombre"]] if col_map.get("nombre") is not None else None
        )
        if not nombre_raw or str(nombre_raw).strip() == "":
            continue

        nombre = str(nombre_raw).strip().upper()

        # Dedup dentro del mismo Excel
        if nombre in seen_names:
            warnings.append(f"Empresa duplicada en Excel (ignorada): {nombre}")
            continue
        seen_names.add(nombre)

        # Ciudades
        ciudades: list[str] = []
        for abrev, ci in ciudad_cols.items():
            val = row[ci] if ci < len(row) else None
            if val and str(val).strip().upper() in (
                "X",
                "SI",
                "SÍ",
                "YES",
                "1",
                "TRUE",
            ):
                cn = CIUDAD_ABREV.get(abrev.upper(), abrev.upper())
                if cn not in ciudades:
                    ciudades.append(cn)

        if ciudades_csv_col is not None and not ciudades:
            csv_val = row[ciudades_csv_col] if ciudades_csv_col < len(row) else None
            if csv_val:
                for c in str(csv_val).split(","):
                    cn = CIUDAD_ABREV.get(c.strip().upper(), c.strip().upper())
                    if cn and cn not in ciudades:
                        ciudades.append(cn)

        if not ciudades:
            ciudades = ["MADRID"]

        def _get(field, default=None):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        # ── Detect format: NEW (EF/IT split) vs OLD (single frecuenciaSolicitada) ──
        # New format: has frecuenciaEF and/or frecuenciaIT columns
        # Old format: has frecuenciaSolicitada column only
        has_ef_col = "frecuenciaEF" in col_map
        has_it_col = "frecuenciaIT" in col_map
        has_old_col = "frecuenciaSolicitada" in col_map

        freq_ef_raw = _get("frecuenciaEF") if has_ef_col else None
        freq_it_raw = _get("frecuenciaIT") if has_it_col else None
        freq_old_raw = _get("frecuenciaSolicitada") if has_old_col else None

        # Parse frequencies
        freq_ef = _int(freq_ef_raw) if freq_ef_raw not in (None, "", 0) else None
        freq_it = _int(freq_it_raw) if freq_it_raw not in (None, "", 0) else None
        freq_total = None

        if has_ef_col or has_it_col:
            # NEW format: use EF/IT values
            ef_val = freq_ef or 0
            it_val = freq_it or 0
            freq_total = ef_val + it_val if (ef_val > 0 or it_val > 0) else None
        elif freq_old_raw not in (None, "", 0):
            # OLD format: use frecuenciaSolicitada as total
            freq_total = _int(freq_old_raw)
            # Split proportionally based on tipo
            tipo_val = _programa(_get("tipo"))
            if tipo_val == "EF":
                freq_ef = freq_total
                freq_it = 0
            elif tipo_val == "IT":
                freq_ef = 0
                freq_it = freq_total
            else:  # AMBAS — split evenly
                freq_ef = freq_total // 2
                freq_it = freq_total - freq_ef

        # ── ConfigTrimestral fields wired in V18 ──────────────
        # None = column absent or cell blank → keep DB default on UPDATE,
        # fall back to legacy literal on INSERT.
        escuela_propia_raw = _get("escuelaPropia")
        escuela_propia_val = (
            _bool(escuela_propia_raw) if escuela_propia_raw is not None else None
        )

        disponibilidad_raw = _get("disponibilidadDias")
        disponibilidad_val = (
            _normalizar_dias(disponibilidad_raw) if disponibilidad_raw else None
        )

        voluntarios_val = _to_int_or_none(_get("voluntariosDisponibles"))

        # NOTE: scoreV3, semaforo, fiabilidadReciente are IGNORED from Excel
        # These fields are now auto-calculated from historical data
        empresas_data.append(
            {
                "nombre": nombre,
                "tipo": _programa(_get("tipo")),
                # scoreV3, semaforo, fiabilidadReciente omitted — auto-calculated
                "esComodin": _bool(_get("esComodin")),
                "aceptaExtras": _bool(_get("aceptaExtras")),
                "maxExtrasTrimestre": _int(_get("maxExtrasTrimestre")),
                "prioridadReduccion": _prioridad(_get("prioridadReduccion")),
                "tieneBolsa": _bool(_get("tieneBolsa")),
                "turnoPreferido": _str(_get("turnoPreferido")),
                "activa": _bool(_get("activa", "SI")),
                "esNueva": _bool(_get("esNueva", "NO")),
                "notas": _str(_get("notas")),
                "ciudades": ciudades,
                # Frequency fields
                "frecuenciaSolicitada": freq_total,
                "frecuenciaEF": freq_ef,
                "frecuenciaIT": freq_it,
                # ConfigTrimestral fields (V18) — None means "not provided"
                "escuelaPropia": escuela_propia_val,
                "disponibilidadDias": disponibilidad_val,
                "voluntariosDisponibles": voluntarios_val,
            }
        )

    if not empresas_data:
        raise HTTPException(400, "No se encontraron empresas en el Excel")

    # ── Cargar empresas existentes (CASE-INSENSITIVE) ────────
    existing_rows = await db.execute(text("SELECT id, nombre FROM empresa"))
    db_lookup: dict[str, tuple[int, str]] = {}
    for r in existing_rows.mappings().all():
        db_lookup[r["nombre"].strip().upper()] = (r["id"], r["nombre"])

    # ── Cargar ciudades existentes (CASE-INSENSITIVE) ────────
    existing_cities = await db.execute(text("SELECT id, nombre FROM ciudad"))
    city_lookup: dict[str, int] = {}
    for r in existing_cities.mappings().all():
        city_lookup[r["nombre"].strip().upper()] = r["id"]

    # ── 1. Crear ciudades nuevas ─────────────────────────────
    all_cities = set()
    for emp in empresas_data:
        all_cities.update(c.upper() for c in emp["ciudades"])

    ciudades_creadas = []
    for cn in sorted(all_cities):
        if cn not in city_lookup:
            r = await db.execute(
                text(
                    "INSERT INTO ciudad (nombre, activa) VALUES (:n, true) RETURNING id"
                ),
                {"n": cn},
            )
            city_lookup[cn] = r.scalar_one()
            ciudades_creadas.append(cn)

    # ── 2. Upsert empresas ───────────────────────────────────
    creadas = 0
    actualizadas = 0
    eid_map: dict[str, int] = {}

    for emp in empresas_data:
        key = emp["nombre"].upper()
        existing = db_lookup.get(key)

        if existing:
            eid, old_name = existing
            # NOTE: scoreV3, semaforo, fiabilidadReciente NOT updated — auto-calculated
            await db.execute(
                text("""
                    UPDATE empresa SET
                        nombre = :nombre, tipo = :tipo,
                        "esComodin" = :esComodin, "aceptaExtras" = :aceptaExtras,
                        "maxExtrasTrimestre" = :maxExtrasTrimestre,
                        "prioridadReduccion" = :prioridadReduccion,
                        "tieneBolsa" = :tieneBolsa, "turnoPreferido" = :turnoPreferido,
                        activa = :activa, "esNueva" = :esNueva,
                        notas = :notas, "updatedAt" = NOW()
                    WHERE id = :id
                """),
                {**emp, "id": eid},
            )
            actualizadas += 1
            eid_map[key] = eid
            if old_name != emp["nombre"]:
                warnings.append(f"'{old_name}' → '{emp['nombre']}'")
        else:
            # New empresa: set neutral scores (will be calculated after first quarter)
            r = await db.execute(
                text("""
                    INSERT INTO empresa (
                        nombre, tipo, semaforo, "scoreV3", "fiabilidadReciente",
                        "esComodin", "aceptaExtras", "maxExtrasTrimestre",
                        "prioridadReduccion", "tieneBolsa", "turnoPreferido",
                        activa, "esNueva", notas, "updatedAt"
                    ) VALUES (
                        :nombre, :tipo, 'AMBAR', 50, 50,
                        :esComodin, :aceptaExtras, :maxExtrasTrimestre,
                        :prioridadReduccion, :tieneBolsa, :turnoPreferido,
                        :activa, :esNueva, :notas, NOW()
                    ) RETURNING id
                """),
                emp,
            )
            eid = r.scalar_one()
            creadas += 1
            eid_map[key] = eid
            db_lookup[key] = (eid, emp["nombre"])

    # ── 3. Sync empresaCiudad ────────────────────────────────
    ec_count = 0
    for emp in empresas_data:
        eid = eid_map[emp["nombre"].upper()]
        await db.execute(
            text('DELETE FROM "empresaCiudad" WHERE "empresaId" = :eid'),
            {"eid": eid},
        )
        for cn in emp["ciudades"]:
            cid = city_lookup.get(cn.upper())
            if cid:
                await db.execute(
                    text("""
                        INSERT INTO "empresaCiudad" ("empresaId", "ciudadId", "activaReciente")
                        VALUES (:eid, :cid, true)
                        ON CONFLICT ("empresaId", "ciudadId") DO UPDATE SET "activaReciente" = true
                    """),
                    {"eid": eid, "cid": cid},
                )
                ec_count += 1

    # ── 4. Upsert configTrimestral ───────────────────────────
    cfg_created = 0
    cfg_updated = 0
    for emp in empresas_data:
        if not emp["activa"]:
            continue
        eid = eid_map[emp["nombre"].upper()]

        # Determine tipoParticipacion from EF/IT frequencies
        freq_ef = emp.get("frecuenciaEF") or 0
        freq_it = emp.get("frecuenciaIT") or 0
        if freq_ef > 0 and freq_it > 0:
            tipo_calc = "AMBAS"
        elif freq_ef > 0:
            tipo_calc = "EF"
        elif freq_it > 0:
            tipo_calc = "IT"
        else:
            tipo_calc = emp["tipo"]  # Fallback to empresa.tipo

        # Check if config exists
        existing = await db.execute(
            text("""
                SELECT id FROM "configTrimestral"
                WHERE trimestre = :tri AND "empresaId" = :eid
            """),
            {"tri": trimestre, "eid": eid},
        )

        if existing.first():
            # Update existing config (preserve other fields, update freq/tipo/notas)
            update_fragments = [
                '"tipoParticipacion" = :tipo',
                '"turnoPreferido" = :turno',
                '"frecuenciaSolicitada" = :freq',
                '"frecuenciaEF" = :freq_ef',
                '"frecuenciaIT" = :freq_it',
                '"notas" = COALESCE(:notas, "notas")',
                '"updatedAt" = NOW()',
            ]
            update_params = {
                "eid": eid,
                "tri": trimestre,
                "tipo": tipo_calc,
                "turno": emp["turnoPreferido"],
                "freq": emp.get("frecuenciaSolicitada"),
                "freq_ef": emp.get("frecuenciaEF"),
                "freq_it": emp.get("frecuenciaIT"),
                "notas": emp.get("notas"),
            }

            # V18: only override these when Excel explicitly provided a value
            if emp.get("escuelaPropia") is not None:
                update_fragments.append('"escuelaPropia" = :escuela')
                update_params["escuela"] = emp["escuelaPropia"]

            if emp.get("disponibilidadDias") is not None:
                update_fragments.append('"disponibilidadDias" = :dias')
                update_params["dias"] = emp["disponibilidadDias"]

            if emp.get("voluntariosDisponibles") is not None:
                update_fragments.append('"voluntariosDisponibles" = :vol')
                update_params["vol"] = emp["voluntariosDisponibles"]

            await db.execute(
                text(f"""
                    UPDATE "configTrimestral"
                    SET {', '.join(update_fragments)}
                    WHERE trimestre = :tri AND "empresaId" = :eid
                """),
                update_params,
            )
            cfg_updated += 1
        else:
            # Create new config
            # V18: bind escuelaPropia / disponibilidadDias / voluntariosDisponibles,
            # falling back to legacy defaults when the Excel did not provide them.
            escuela_val = (
                emp["escuelaPropia"] if emp.get("escuelaPropia") is not None else False
            )
            dias_val = (
                emp["disponibilidadDias"]
                if emp.get("disponibilidadDias") is not None
                else "L,M,X,J,V"
            )
            vol_val = (
                emp["voluntariosDisponibles"]
                if emp.get("voluntariosDisponibles") is not None
                else 0
            )

            await db.execute(
                text("""
                    INSERT INTO "configTrimestral" (
                        "empresaId", trimestre, "tipoParticipacion",
                        "escuelaPropia", "disponibilidadDias", "turnoPreferido",
                        "frecuenciaSolicitada", "frecuenciaEF", "frecuenciaIT",
                        "voluntariosDisponibles", "notas", "createdAt", "updatedAt"
                    ) VALUES (
                        :eid, :tri, :tipo, :escuela, :dias, :turno,
                        :freq, :freq_ef, :freq_it, :vol, :notas, NOW(), NOW()
                    )
                """),
                {
                    "eid": eid,
                    "tri": trimestre,
                    "tipo": tipo_calc,
                    "escuela": escuela_val,
                    "dias": dias_val,
                    "turno": emp["turnoPreferido"],
                    "freq": emp.get("frecuenciaSolicitada"),
                    "freq_ef": emp.get("frecuenciaEF"),
                    "freq_it": emp.get("frecuenciaIT"),
                    "vol": vol_val,
                    "notas": emp.get("notas"),
                },
            )
            cfg_created += 1

    # ── 5. Festivos (dates-based) ──────────────────────────────
    from datetime import datetime, date, timedelta

    festivo_count = 0
    festivo_sheet = None
    # Accept multiple sheet name variants
    for c in ["festivos", "festivo", "semanasexcluidas", "semanas_excluidas", "semanas excluidas"]:
        if c in sheet_map:
            festivo_sheet = wb[sheet_map[c]]
            break

    if festivo_sheet:
        # Helper: convert fecha to festivo fields
        DIA_LETRA = {0: "L", 1: "M", 2: "X", 3: "J", 4: "V", 5: "S", 6: "D"}
        Q_MONTH = {1: "Q1", 2: "Q1", 3: "Q1", 4: "Q2", 5: "Q2", 6: "Q2",
                   7: "Q3", 8: "Q3", 9: "Q3", 10: "Q4", 11: "Q4", 12: "Q4"}

        def fecha_to_festivo_fields(fecha: date) -> dict:
            """Given a date, returns {dia, trimestre, semana} for the festivo table."""
            dia = DIA_LETRA[fecha.weekday()]
            year = fecha.year
            tri = f"{year}-{Q_MONTH[fecha.month]}"

            # Calculate relative week: find first Monday of the quarter
            quarter_num = (fecha.month - 1) // 3 + 1
            quarter_start_month = (quarter_num - 1) * 3 + 1
            first_day = date(year, quarter_start_month, 1)

            if first_day.weekday() == 0:
                first_monday = first_day
            else:
                days_until_monday = (7 - first_day.weekday()) % 7
                first_monday = first_day + timedelta(days=days_until_monday)

            days_diff = (fecha - first_monday).days
            semana = (days_diff // 7) + 1
            semana = max(1, min(13, semana))

            return {"dia": dia, "trimestre": tri, "semana": semana}

        festivos_to_insert: list[dict] = []

        # Check header format (new: Fecha, Motivo | old: Trimestre, Semana, Motivo)
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in festivo_sheet[1]]
        is_new_format = "fecha" in headers

        for row_num, row in enumerate(festivo_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue

            if is_new_format:
                # New format: Fecha | Motivo
                fecha_val = row[0]
                motivo = _str(row[1]) if len(row) > 1 and row[1] else None

                # Parse fecha
                fecha = None
                if isinstance(fecha_val, datetime):
                    fecha = fecha_val.date()
                elif isinstance(fecha_val, date):
                    fecha = fecha_val
                elif isinstance(fecha_val, str):
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                        try:
                            fecha = datetime.strptime(fecha_val.strip(), fmt).date()
                            break
                        except ValueError:
                            continue

                if not fecha:
                    warnings.append(f"Festivos fila {row_num}: Fecha inválida '{fecha_val}'")
                    continue

                # Skip weekends
                if fecha.weekday() >= 5:
                    warnings.append(f"Festivos fila {row_num}: {fecha} cae en fin de semana — ignorado")
                    continue

                fields = fecha_to_festivo_fields(fecha)
                festivos_to_insert.append({
                    "fecha": fecha,
                    "dia": fields["dia"],
                    "trimestre": fields["trimestre"],
                    "semana": fields["semana"],
                    "motivo": motivo,
                })
            else:
                # Legacy format: Trimestre | Semana | Motivo
                # Convert to dates using approximate calculation (for backward compat)
                tri = _str(row[0]) if len(row) > 0 else None
                sem = _int(row[1]) if len(row) > 1 else 0
                motivo = _str(row[2]) if len(row) > 2 else None
                if not tri or sem <= 0:
                    continue

                # Legacy: exclude all 5 days of the week
                year = int(tri[:4])
                quarter = int(tri[-1])
                quarter_start_month = {1: 1, 2: 4, 3: 7, 4: 10}[quarter]
                first_day = date(year, quarter_start_month, 1)

                if first_day.weekday() == 0:
                    first_monday = first_day
                else:
                    days_until_monday = (7 - first_day.weekday()) % 7
                    first_monday = first_day + timedelta(days=days_until_monday)

                week_start = first_monday + timedelta(weeks=sem - 1)

                # Add all 5 weekdays
                for day_offset in range(5):
                    fecha = week_start + timedelta(days=day_offset)
                    dia = DIA_LETRA[fecha.weekday()]
                    festivos_to_insert.append({
                        "fecha": fecha,
                        "dia": dia,
                        "trimestre": tri,
                        "semana": sem,
                        "motivo": motivo,
                    })

        # Delete existing festivos for the year(s) being imported
        years = set(f["fecha"].year for f in festivos_to_insert)
        for year in years:
            await db.execute(
                text("DELETE FROM festivo WHERE EXTRACT(YEAR FROM fecha) = :year"),
                {"year": year},
            )

        # Insert new festivos
        for f in festivos_to_insert:
            await db.execute(
                text("""
                    INSERT INTO festivo (fecha, dia, trimestre, semana, motivo, "createdAt")
                    VALUES (:fecha, :dia, :trimestre, :semana, :motivo, NOW())
                    ON CONFLICT (fecha) DO UPDATE SET
                        dia = EXCLUDED.dia,
                        trimestre = EXCLUDED.trimestre,
                        semana = EXCLUDED.semana,
                        motivo = EXCLUDED.motivo
                """),
                f,
            )
            festivo_count += 1

    await db.commit()

    return ImportEmpresasResult(
        total_empresas=len(empresas_data),
        creadas=creadas,
        actualizadas=actualizadas,
        ciudades_creadas=ciudades_creadas,
        empresa_ciudad_links=ec_count,
        config_trimestral_creadas=cfg_created,
        configs_actualizadas=cfg_updated,
        trimestre_target=trimestre,
        festivos_importados=festivo_count,
        warnings=warnings,
    )


# ══════════════════════════════════════════════════════════════
# ENDPOINT 2: Importar histórico (unificado)
# ══════════════════════════════════════════════════════════════
 
 
@router.post("/historico", response_model=ImportHistoricoResult)
async def importar_historico(
    file: UploadFile = File(...),
    trimestre: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Importa un calendario completado como histórico del trimestre.
 
    Acepta el Excel exportado por /exportar-excel (formato estándar):
        Semana | Día | Horario | Turno | Empresa | Taller | Programa | Ciudad | Tipo | Estado | Confirmado
 
    También acepta formatos legacy con columnas:
        SEM | D | TALLER | EMPRESA | CIUDAD | FECHA | TIPO
 
    Lógica de Estado:
    - OK / PLANIFICADO / SÍ → se importa como 'OK'
    - CANCELADO / NO → se importa como 'CANCELADO'
    - VACANTE sin empresa → se ignora
    - VACANTE con empresa (rellenada manualmente) → se importa como 'OK'
 
    Idempotente: borra histórico anterior del mismo trimestre.
    """
    import openpyxl
    from datetime import date, timedelta, datetime
 
    content = await file.read()
    filename = (file.filename or "").lower()
 
    # Soportar CSV legacy también
    if filename.endswith(".csv"):
        import pandas as pd
        df = pd.read_csv(BytesIO(content))
        # Convertir a openpyxl workbook en memoria
        wb = openpyxl.Workbook()
        ws_tmp = wb.active
        ws_tmp.append(list(df.columns))
        for _, row in df.iterrows():
            ws_tmp.append(list(row))
    else:
        wb = openpyxl.load_workbook(BytesIO(content))
 
    ws = wb.active
    warnings: list[str] = []
 
    # ── Detectar columnas (flexible, soporta ambos formatos) ──
    raw_headers = [str(c.value or "").strip() for c in ws[1]]
    headers_lower = [h.lower().replace("í", "i").replace("á", "a") for h in raw_headers]
 
    COLUMN_ALIASES = {
        "semana": ["semana", "sem", "week"],
        "dia": ["dia", "día", "d", "day"],
        "horario": ["horario", "hora", "time"],
        "turno": ["turno", "shift"],
        "empresa": ["empresa", "company", "emp"],
        "empresa_original": ["empresa original", "empresa_original", "empresaoriginal", "original company"],
        "taller": ["taller", "workshop", "tal"],
        "programa": ["programa", "prog", "tipo", "type"],
        "ciudad": ["ciudad", "city"],
        "estado": ["estado", "status"],
        "fecha": ["fecha", "date", "fecha_taller"],
        "motivo_cambio": ["motivo cambio", "motivo_cambio", "motivocambio", "motivo"],
    }
 
    col_idx: dict[str, int] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            for i, h in enumerate(headers_lower):
                if h == alias and field not in col_idx:
                    col_idx[field] = i
                    break
            if field in col_idx:
                break
 
    # Validar mínimo
    if "empresa" not in col_idx or "taller" not in col_idx:
        raise HTTPException(
            400,
            f"Columnas requeridas 'Empresa' y 'Taller' no encontradas. "
            f"Headers: {raw_headers}",
        )
 
    has_semana = "semana" in col_idx
    has_fecha = "fecha" in col_idx
    has_estado = "estado" in col_idx
 
    if not has_semana and not has_fecha:
        warnings.append(
            "No se encontró columna 'Semana' ni 'Fecha'. "
            "Se usará fecha genérica para todas las filas."
        )
 
    # ── Lookups de BD ────────────────────────────────────────
    emp_rows = await db.execute(text("SELECT id, nombre FROM empresa"))
    emp_map = {
        r["nombre"].strip().upper(): r["id"]
        for r in emp_rows.mappings().all()
    }
 
    taller_rows = await db.execute(
        text("SELECT id, nombre FROM taller WHERE activo = true")
    )
    talleres = [dict(r) for r in taller_rows.mappings().all()]
 
    def match_taller(n: str) -> int | None:
        nl = n.strip().lower()
        for t in talleres:
            if t["nombre"].strip().lower() == nl:
                return t["id"]
        for t in talleres:
            if nl in t["nombre"].strip().lower() or t["nombre"].strip().lower() in nl:
                return t["id"]
        return None
 
    # ── Función para calcular fecha real ─────────────────────
    def fecha_from_semana(sem_rel: int, dia: str) -> date:
        year = int(trimestre.split("-")[0])
        q = trimestre.split("-")[1]
        q_start_month = {"Q1": 1, "Q2": 4, "Q3": 7, "Q4": 10}
        month = q_start_month.get(q, 1)
        first_day = date(year, month, 1)
        days_to_monday = (7 - first_day.weekday()) % 7
        first_monday = first_day + timedelta(days=days_to_monday)
        if first_day.weekday() == 0:
            first_monday = first_day
        target_monday = first_monday + timedelta(weeks=sem_rel - 1)
        DIA_OFFSET = {"L": 0, "M": 1, "X": 2, "J": 3, "V": 4}
        return target_monday + timedelta(days=DIA_OFFSET.get(dia, 0))
 
    def fecha_fallback() -> date:
        """Fecha genérica si no hay semana ni fecha."""
        year = int(trimestre.split("-")[0])
        q = trimestre.split("-")[1]
        q_start_month = {"Q1": 1, "Q2": 4, "Q3": 7, "Q4": 10}
        return date(year, q_start_month.get(q, 1), 1)
 
    # Normalización legacy (del importar.py original)
    NORMALIZE_EMPRESA = {
        "ACCIONA ": "ACCIONA",
        "AECOM ": "AECOM",
        "ATREVIA ": "ATREVIA",
        " INDRA": "INDRA",
        "MP CAPGEMINI": "CAPGEMINI",
        "MP ENDESA": "ENDESA",
        "MP LDA": "LDA",
        "MP AON": "F. AON",
        "MP SANTANDER": "BANCO SANTANDER",
        "SERVITEC / CAPGEMINI": "SERVITEC",
        "SERVITEC / SERVITEC": "SERVITEC",
        "URBASER / SACYR": "URBASER",
    }
 
    # ── Borrar histórico anterior ────────────────────────────
    await db.execute(
        text('DELETE FROM "historicoTaller" WHERE trimestre = :tri'),
        {"tri": trimestre},
    )
 
    # ── Procesar filas ───────────────────────────────────────
    emp_404: set[str] = set()
    taller_404: set[str] = set()
    importadas_ok = 0
    importadas_cancelado = 0
    vacantes_ignoradas = 0
    total_filas = 0
 
    def _cell(row_vals, field: str, default=None):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row_vals):
            return default
        return row_vals[idx]
 
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        total_filas += 1
 
        empresa_raw = _cell(row, "empresa")
        taller_raw = _cell(row, "taller")
 
        # Ignorar filas sin empresa ni taller
        if not empresa_raw or str(empresa_raw).strip() == "":
            vacantes_ignoradas += 1
            continue
        if not taller_raw or str(taller_raw).strip() == "":
            continue
 
        # Estado
        if has_estado:
            estado_raw = str(_cell(row, "estado", "OK")).strip().upper()
        else:
            estado_raw = "OK"  # Legacy sin columna estado → todo OK
 
        if estado_raw in ("VACANTE",) and str(empresa_raw).strip() == "":
            vacantes_ignoradas += 1
            continue
 
        # Normalizar estado
        if estado_raw in ("OK", "PLANIFICADO", "SI", "SÍ", "YES", "1", "VACANTE"):
            # VACANTE con empresa → fue rellenada manualmente → OK
            estado_db = "OK"
        elif estado_raw in ("CANCELADO", "NO", "CANCEL"):
            estado_db = "CANCELADO"
        else:
            estado_db = "OK"
 
        # Resolver empresa (final)
        empresa_nombre = str(empresa_raw).strip()
        empresa_upper = empresa_nombre.upper()
        # Intentar normalización legacy
        empresa_upper = NORMALIZE_EMPRESA.get(empresa_nombre, empresa_upper)
        eid = emp_map.get(empresa_upper)
        if not eid:
            emp_404.add(empresa_upper)
            continue

        # Resolver empresa original (optional — defaults to eid if not present)
        eid_original = eid  # Default: same as final empresa
        empresa_orig_raw = _cell(row, "empresa_original")
        if empresa_orig_raw and str(empresa_orig_raw).strip():
            empresa_orig_nombre = str(empresa_orig_raw).strip()
            empresa_orig_upper = empresa_orig_nombre.upper()
            empresa_orig_upper = NORMALIZE_EMPRESA.get(empresa_orig_nombre, empresa_orig_upper)
            eid_orig_lookup = emp_map.get(empresa_orig_upper)
            if eid_orig_lookup:
                eid_original = eid_orig_lookup
            # If not found, just use eid (don't fail — legacy data may not have this)

        # Resolver motivo_cambio (optional)
        motivo_cambio = None
        motivo_raw = _cell(row, "motivo_cambio")
        if motivo_raw and str(motivo_raw).strip():
            motivo_str = str(motivo_raw).strip().upper().replace(" ", "_")
            # Normalize common variants
            if motivo_str in ("EMPRESA_CANCELO", "EMPRESA CANCELO", "EMPRESA CANCELÓ"):
                motivo_cambio = "EMPRESA_CANCELO"
            elif motivo_str in ("DECISION_PLANIFICADOR", "DECISION PLANIFICADOR", "DECISIÓN PLANIFICADOR"):
                motivo_cambio = "DECISION_PLANIFICADOR"
            # Display text from Excel export
            elif "CANCELÓ" in motivo_str or "CANCELO" in motivo_str:
                motivo_cambio = "EMPRESA_CANCELO"
            elif "PLANIFICADOR" in motivo_str:
                motivo_cambio = "DECISION_PLANIFICADOR"

        # Resolver taller
        tid = match_taller(str(taller_raw).strip())
        if not tid:
            taller_404.add(str(taller_raw).strip())
            continue

        # Resolver fecha
        fecha = None
        if has_fecha:
            fecha_raw = _cell(row, "fecha")
            if isinstance(fecha_raw, datetime):
                fecha = fecha_raw.date()
            elif isinstance(fecha_raw, date):
                fecha = fecha_raw

        if fecha is None and has_semana:
            try:
                sem = int(float(_cell(row, "semana", 1)))
                dia = str(_cell(row, "dia", "L")).strip().upper()
                fecha = fecha_from_semana(sem, dia)
            except Exception:
                fecha = fecha_fallback()

        if fecha is None:
            fecha = fecha_fallback()

        # Ciudad
        ciudad = str(_cell(row, "ciudad", "MADRID")).strip()
        if not ciudad:
            ciudad = "MADRID"

        # Insertar
        await db.execute(
            text("""
                INSERT INTO "historicoTaller"
                    ("empresaId", "empresaIdOriginal", "tallerId", fecha, estado, ciudad, trimestre, "motivoCambio")
                VALUES (:eid, :eid_original, :tid, :fecha, :estado, :ciudad, :tri, :motivo)
            """),
            {
                "eid": eid,
                "eid_original": eid_original,
                "tid": tid,
                "fecha": fecha,
                "estado": estado_db,
                "ciudad": ciudad,
                "tri": trimestre,
                "motivo": motivo_cambio,
            },
        )
 
        if estado_db == "OK":
            importadas_ok += 1
        else:
            importadas_cancelado += 1
 
    await db.commit()
 
    if emp_404:
        warnings.append(f"Empresas no encontradas: {sorted(emp_404)}")
    if taller_404:
        warnings.append(f"Talleres no encontrados: {sorted(taller_404)}")
 
    return ImportHistoricoResult(
        trimestre=trimestre,
        total_filas=total_filas,
        importadas_ok=importadas_ok,
        importadas_cancelado=importadas_cancelado,
        vacantes_ignoradas=vacantes_ignoradas,
        empresas_no_encontradas=sorted(emp_404),
        talleres_no_encontrados=sorted(taller_404),
        warnings=warnings,
    )


# ══════════════════════════════════════════════════════════════
# ENDPOINT 3: Estado
# ══════════════════════════════════════════════════════════════


@router.get("/estado/{trimestre}")
async def estado_importacion(trimestre: str, db: AsyncSession = Depends(get_db)):
    counts = {}
    for key, q in [
        ("empresas_activas", "SELECT COUNT(*) FROM empresa WHERE activa = true"),
        (
            "config_trimestral",
            'SELECT COUNT(*) FROM "configTrimestral" WHERE trimestre = :tri',
        ),
        ("frecuencias", "SELECT COUNT(*) FROM frecuencia WHERE trimestre = :tri"),
        ("historico", 'SELECT COUNT(*) FROM "historicoTaller" WHERE trimestre = :tri'),
        ("restricciones", "SELECT COUNT(*) FROM restriccion"),
        ("ciudades", "SELECT COUNT(*) FROM ciudad"),
        ("empresa_ciudad", 'SELECT COUNT(*) FROM "empresaCiudad"'),
    ]:
        r = await db.execute(text(q), {"tri": trimestre} if ":tri" in q else {})
        counts[key] = r.scalar()

    try:
        r = await db.execute(
            text('SELECT COUNT(*) FROM festivo WHERE trimestre = :tri'),
            {"tri": trimestre},
        )
        counts["festivos"] = r.scalar()
    except:  # noqa: E722
        counts["festivos"] = 0

    return {
        "trimestre": trimestre,
        **counts,
        "listo_fase_1": counts["config_trimestral"] > 0,
        "listo_fase_2": counts["frecuencias"] > 0,
    }


# ══════════════════════════════════════════════════════════════
# ENDPOINT 4: Clonar trimestre
# ══════════════════════════════════════════════════════════════

class ClonarTrimestreResult(BaseModel):
    trimestre_origen: str
    trimestre_destino: str
    configs_clonadas: int
    configs_ya_existentes: int
    empresas_saltadas: list[str]   # inactivas, no se clonan
    warnings: list[str]


@router.post("/clonar-trimestre", response_model=ClonarTrimestreResult)
async def clonar_trimestre(
    trimestre_origen: str = Form(...),   # "2026-Q2"
    trimestre_destino: str = Form(...),  # "2026-Q3"
    db: AsyncSession = Depends(get_db),
):
    """
    Clona las configTrimestral activas de un trimestre al siguiente.

    Copia: tipoParticipacion, turnoPreferido, frecuenciaSolicitada,
           disponibilidadDias.
    NO copia: frecuencias calculadas, planificaciones, semanas excluidas
              (las semanas excluidas ya deben estar cargadas para todo el año).

    Comportamiento:
    - Si ya existe config para (empresa, destino) → NO sobreescribe (ON CONFLICT DO NOTHING)
    - Empresas inactivas (activa=false) → se omiten y se listan en empresas_saltadas
    """
    warnings: list[str] = []

    # ── Leer configs del trimestre origen ───────────────────
    rows = await db.execute(
        text("""
            SELECT
                ct."empresaId",
                e.nombre,
                e.activa,
                ct."tipoParticipacion",
                ct."turnoPreferido",
                ct."frecuenciaSolicitada",
                ct."disponibilidadDias"
            FROM "configTrimestral" ct
            JOIN empresa e ON e.id = ct."empresaId"
            WHERE ct.trimestre = :origen
            ORDER BY e.nombre
        """),
        {"origen": trimestre_origen},
    )
    configs = [dict(r) for r in rows.mappings().all()]

    if not configs:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=404,
            detail=f"No hay configuraciones para el trimestre {trimestre_origen}. "
                   "Importa primero ese trimestre.",
        )

    clonadas = 0
    ya_existentes = 0
    saltadas: list[str] = []

    for cfg in configs:
        # Saltar empresas dadas de baja definitivamente
        if not cfg["activa"]:
            saltadas.append(cfg["nombre"])
            continue

        result = await db.execute(
            text("""
                INSERT INTO "configTrimestral" (
                    "empresaId", trimestre, "tipoParticipacion",
                    "escuelaPropia", "disponibilidadDias",
                    "turnoPreferido", "frecuenciaSolicitada", "updatedAt"
                )
                VALUES (
                    :eid, :destino, :tipo,
                    false, :dias,
                    :turno, :freq, NOW()
                )
                ON CONFLICT ("empresaId", trimestre) DO NOTHING
            """),
            {
                "eid": cfg["empresaId"],
                "destino": trimestre_destino,
                "tipo": cfg["tipoParticipacion"],
                "dias": cfg["disponibilidadDias"] or "L,M,X,J,V",
                "turno": cfg["turnoPreferido"],
                "freq": cfg["frecuenciaSolicitada"],
            },
        )
        # rowcount = 0 → ya existía (DO NOTHING), 1 → insertado
        if result.rowcount == 0:
            ya_existentes += 1
        else:
            clonadas += 1

    await db.commit()

    if saltadas:
        warnings.append(
            f"{len(saltadas)} empresa(s) inactiva(s) no clonadas: {', '.join(saltadas)}"
        )
    if ya_existentes > 0:
        warnings.append(
            f"{ya_existentes} config(s) ya existían en {trimestre_destino} y no se modificaron."
        )

    return ClonarTrimestreResult(
        trimestre_origen=trimestre_origen,
        trimestre_destino=trimestre_destino,
        configs_clonadas=clonadas,
        configs_ya_existentes=ya_existentes,
        empresas_saltadas=saltadas,
        warnings=warnings,
    )


# ══════════════════════════════════════════════════════════════
# ENDPOINT 5: Consultar festivos
# ══════════════════════════════════════════════════════════════


class FestivoOut(BaseModel):
    id: int
    fecha: str
    dia: str
    trimestre: str
    semana: int
    motivo: str | None


class FestivosResult(BaseModel):
    year: int
    total: int
    festivos: list[FestivoOut]


@router.get("/festivos/{year}", response_model=FestivosResult)
async def obtener_festivos(year: int, db: AsyncSession = Depends(get_db)):
    """Returns all festivos for a given year."""
    result = await db.execute(
        text("""
            SELECT id, fecha, dia, trimestre, semana, motivo
            FROM festivo
            WHERE EXTRACT(YEAR FROM fecha) = :year
            ORDER BY fecha
        """),
        {"year": year},
    )
    rows = [dict(r) for r in result.mappings().all()]

    # Convert date to string for JSON serialization
    festivos = []
    for row in rows:
        festivos.append({
            "id": row["id"],
            "fecha": row["fecha"].strftime("%Y-%m-%d") if row["fecha"] else "",
            "dia": row["dia"],
            "trimestre": row["trimestre"],
            "semana": row["semana"],
            "motivo": row["motivo"],
        })

    return FestivosResult(year=year, total=len(festivos), festivos=festivos)

