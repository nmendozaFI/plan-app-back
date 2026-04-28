"""
FASE 2 — Solver de Calendario (OR-Tools CP-SAT) v3
MODELO DE SLOTS FIJOS: 20 talleres fijos por semana (14 EF + 6 IT).

Cambio fundamental respecto a v2:
  - Cada taller tiene día + horario fijo (no flexible)
  - Cada semana repite los mismos 20 slots
  - El solver decide QUÉ EMPRESA va a cada slot de cada semana
  - Post-proceso asigna ciudades (talleres ya están fijos)

Variables de decisión:
  assign[empresa, semana, slot] = 1 si empresa ocupa ese slot esa semana

Tablas Prisma (camelCase):
  - frecuencia, planificacion, restriccion, taller,
    "empresaCiudad", "solverLog"
"""

import asyncio
import logging

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.db import get_db
from app.schemas.calendario import (
    CalendarioInput, CalendarioOutput, SlotCalendario, SugerenciaContingencia,
    SlotUpdateInput, SlotBatchUpdateItem, SlotBatchUpdateInput,
    ValidarAsignacionInput, ValidarAsignacionResult,
    EmpresaAnalisis, CambioSlot, AnalisisResumen, AnalisisResponse,
    CerrarTrimestreInput, CerrarTrimestreResult,
    EmpresaCambiada, CambioDetalle, ImportarExcelResult, ImportarExcelInput,
    ImportarExcelBulkResult, FilaExtraInsertada,
    ListaExtrasResponse, SlotExtraResponse,
    RecalcularScoresResult,
)
from app.services.calendario.solver import (
    _ejecutar_solver,
    _franja_preferida,
    _dias_exclusivos_hard,
)
from app.services.calendario.post_proceso import (
    calcular_fecha_slot, _asignar_ciudades, _guardar_log,
)

router = APIRouter()
logger = logging.getLogger(__name__)


# ── Endpoints ────────────────────────────────────────────────

@router.post("/generar", response_model=CalendarioOutput)
async def generar_calendario(
    params: CalendarioInput,
    db: AsyncSession = Depends(get_db),
):
    """Ejecuta Fase 2: genera el calendario trimestral usando CP-SAT."""
    trimestre = params.trimestre
    warnings: list[str] = []

    # ── 1. Cargar frecuencias confirmadas (output Fase 1) ────
    freq_result = await db.execute(
        text("""
            SELECT f."empresaId", e.nombre,
                   f."talleresEF", f."talleresIT", f."totalAsignado",
                   f."semaforoCalculado", f."scoreCalculado",
                   f."esNueva",
                   e."esComodin", e."turnoPreferido"
            FROM frecuencia f
            JOIN empresa e ON e.id = f."empresaId"
            WHERE f.trimestre = :trimestre
        """),
        {"trimestre": trimestre},
    )
    frecuencias = [dict(r) for r in freq_result.mappings().all()]

    if not frecuencias:
        raise HTTPException(
            status_code=404,
            detail=f"No hay frecuencias confirmadas para {trimestre}. "
                   "Ejecutar y confirmar Fase 1 primero.",
        )

    # ── 2. Restricciones ─────────────────────────────────────
    rest_result = await db.execute(
        text('SELECT "empresaId", tipo, clave, valor, "tallerId" FROM restriccion')
    )
    restricciones = [dict(r) for r in rest_result.mappings().all()]

    # ── 3. Talleres (slots fijos) with per-week calendar overrides ───
    # Uses SemanaConfig to determine which talleres are available each week.
    # In intensive weeks: EF afternoon OFF, IT afternoon moved to Wednesday morning.
    from app.routers.calendario_anual import cargar_talleres_semana

    # Parse trimestre to get year and week range
    anio = int(trimestre.split("-")[0])
    quarter_num = int(trimestre.split("Q")[1])
    iso_week_start = (quarter_num - 1) * 13 + 1

    # Load talleres for each week of the trimestre
    talleres_por_semana: dict[int, list[dict]] = {}
    all_taller_ids: set[int] = set()
    num_semanas = params.semanas  # typically 13

    for semana_rel in range(1, num_semanas + 1):
        iso_week = iso_week_start + semana_rel - 1
        talleres_semana = await cargar_talleres_semana(db, anio, iso_week)
        # Convert to solver format (id, diaSemana, etc.)
        talleres_por_semana[semana_rel] = [
            {
                "id": t["taller_id"],
                "nombre": t["nombre"],
                "programa": t["programa"],
                "diaSemana": t["dia_semana"],
                "horario": t["horario"],
                "turno": t["turno"],
                "es_extra": t.get("es_extra", False),
                "extra_id": t.get("extra_id"),
            }
            for t in talleres_semana
        ]
        for t in talleres_semana:
            all_taller_ids.add(t["taller_id"])

    # For backward compatibility, create a "union" list of all talleres (for summary)
    # Get base talleres info for the union
    base_talleres_result = await db.execute(
        text("""
            SELECT id, nombre, programa, "diaSemana", horario, turno
            FROM taller WHERE activo = true ORDER BY id
        """)
    )
    base_talleres = [dict(r) for r in base_talleres_result.mappings().all()]
    talleres = [t for t in base_talleres if t["id"] in all_taller_ids]

    if len(talleres) == 0:
        raise HTTPException(
            status_code=500,
            detail="No hay talleres activos para este trimestre. "
                   "Verificar configuración de talleres en /planificacion/talleres.",
        )

    # Count by week type for summary
    normal_weeks = sum(1 for s in range(1, num_semanas + 1) if len(talleres_por_semana[s]) >= 18)
    intensive_weeks = num_semanas - normal_weeks
    avg_talleres = sum(len(talleres_por_semana[s]) for s in range(1, num_semanas + 1)) // num_semanas
    talleres_ef = [t for t in talleres if t["programa"] == "EF"]
    talleres_it = [t for t in talleres if t["programa"] == "IT"]
    warnings.append(
        f"Catálogo {trimestre}: {len(talleres_ef)} EF + {len(talleres_it)} IT base, "
        f"{normal_weeks} semanas normales + {intensive_weeks} intensivas, ~{avg_talleres} talleres/semana promedio"
    )

    # ── 4. Ciudad Madrid (todos los talleres son de Madrid) ──
    madrid_result = await db.execute(
        text("SELECT id, nombre FROM ciudad WHERE UPPER(nombre) = 'MADRID' LIMIT 1")
    )
    madrid_row = madrid_result.mappings().first()
    if not madrid_row:
        warnings.append("⚠ Ciudad MADRID no encontrada en BD — ciudadId será null")
        madrid_id = None
    else:
        madrid_id = madrid_row["id"]

    # ── 5. Disponibilidad de días por empresa ────────────────
    disp_result = await db.execute(
        text("""
            SELECT "empresaId", "disponibilidadDias"
            FROM "configTrimestral"
            WHERE trimestre = :trimestre
        """),
        {"trimestre": trimestre},
    )
    disponibilidad_map: dict[int, list[str]] = {}
    for row in disp_result.mappings().all():
        r = dict(row)
        dias_str = r["disponibilidadDias"] or "L,M,X,J,V"
        disponibilidad_map[r["empresaId"]] = [d.strip() for d in dias_str.split(",")]

    # ── 6. Festivos → días excluidos por semana ──────────────────
    # Returns set of (semana_relativa, dia) tuples to exclude specific slots
    dias_excluidos: set[tuple[int, str]] = set()
    semanas_excluidas: set[int] = set()  # Keep for backward compat (weeks where ALL 5 days are excluded)
    try:
        fest_result = await db.execute(
            text('''
                SELECT semana, dia, motivo FROM festivo
                WHERE trimestre = :trimestre
                ORDER BY semana, dia
            '''),
            {"trimestre": trimestre},
        )
        festivos_rows = [dict(r) for r in fest_result.mappings().all()]

        for row in festivos_rows:
            sem = row["semana"]
            dia = row["dia"]
            if 1 <= sem <= 13 and dia in ("L", "M", "X", "J", "V"):
                dias_excluidos.add((sem, dia))

        # Check if any week has ALL 5 days excluded (full week closure like summer)
        from collections import Counter
        dias_por_semana = Counter(sem for sem, _ in dias_excluidos)
        for sem, count in dias_por_semana.items():
            if count >= 5:
                semanas_excluidas.add(sem)

        if dias_excluidos:
            # Format for warning: group by week
            by_week: dict[int, list[str]] = {}
            for sem, dia in sorted(dias_excluidos):
                by_week.setdefault(sem, []).append(dia)
            parts = []
            for sem in sorted(by_week):
                dias = by_week[sem]
                if len(dias) >= 5:
                    parts.append(f"S{sem} (completa)")
                else:
                    parts.append(f"S{sem}:{','.join(dias)}")
            warnings.append(f"Días excluidos: {', '.join(parts)}")

    except Exception:
        pass  # table may not exist yet

    # ── 7. Ejecutar solver (en thread separado para no bloquear event loop) ───
    resultado = await asyncio.to_thread(
        _ejecutar_solver,
        frecuencias=frecuencias,
        restricciones=restricciones,
        talleres=talleres,
        talleres_por_semana=talleres_por_semana,
        disponibilidad_map=disponibilidad_map,
        semanas_excluidas=semanas_excluidas,
        dias_excluidos=dias_excluidos,
        params=params,
    )

    if resultado["status"] in ("INFEASIBLE", "TIMEOUT"):
        await _guardar_log(db, trimestre, resultado)
        await db.commit()
        return CalendarioOutput(**resultado, trimestre=trimestre)

    # ── 8. Post-proceso: asignar ciudad Madrid + sugerencias ──
    slots_completos = _asignar_ciudades(
        slots_raw=resultado["slots"],
        frecuencias=frecuencias,
        madrid_id=madrid_id,
        restricciones=restricciones,
        warnings=warnings,
    )

    resultado["slots"] = slots_completos
    resultado["warnings"] = warnings
    resultado["total_ef"] = sum(1 for s in slots_completos if s["programa"] == "EF" and s["empresa_id"] != 0)
    resultado["total_it"] = sum(1 for s in slots_completos if s["programa"] == "IT" and s["empresa_id"] != 0)
    resultado["total_slots"] = sum(1 for s in slots_completos if s["empresa_id"] != 0)

    # ── 9. Persistir ─────────────────────────────────────────
    await db.execute(
        text('DELETE FROM planificacion WHERE trimestre = :tri'),
        {"tri": trimestre},
    )

    for slot in slots_completos:
        # Vacancies: empresa_id=0 -> store as NULL with estado='VACANTE'
        is_vacancy = slot["empresa_id"] == 0
        empresa_id = None if is_vacancy else slot["empresa_id"]
        estado = "VACANTE" if is_vacancy else "PLANIFICADO"
        tipo_bd = "CONTINGENCIA" if slot["tipo_asignacion"] == "HUECO" else slot["tipo_asignacion"]
        es_contingencia = slot["tipo_asignacion"] == "HUECO"
        await db.execute(
            text("""
                INSERT INTO planificacion
                    (trimestre, semana, dia, horario, turno, "empresaId", "empresaIdOriginal",
                     "tallerId", "ciudadId", "tipoAsignacion", "esContingencia", estado, confirmado, "updatedAt")
                VALUES (
                    :tri, :sem, :dia, :horario, :turno, :eid, :eid_original,
                    :tid, :cid, :tipo, :contingencia, :estado, false, NOW()
                )
            """),
            {
                "tri": trimestre,
                "sem": slot["semana"],
                "dia": slot["dia"],
                "horario": slot["horario"],
                "turno": slot["turno"],
                "eid": empresa_id,
                "eid_original": empresa_id,  # Same as eid at generation time, NEVER updated later
                "tid": slot["taller_id"],
                "cid": slot.get("ciudad_id"),
                "tipo": tipo_bd,
                "contingencia": es_contingencia,
                "estado": estado,
            },
        )

    await _guardar_log(db, trimestre, resultado)
    await db.commit()

    return CalendarioOutput(**resultado, trimestre=trimestre)


@router.get("/{trimestre}")
async def obtener_calendario(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """Lee el calendario generado de un trimestre (incluye vacantes)."""
    result = await db.execute(
        text("""
            SELECT p.id,
                   p.semana, p.dia, p.horario,
                   COALESCE(p.turno, t.turno) AS turno,
                   p."empresaId" AS empresa_id,
                   p."empresaIdOriginal" AS empresa_id_original,
                   e.nombre AS empresa_nombre,
                   t.programa,
                   p."tallerId" AS taller_id,
                   t.nombre AS taller_nombre,
                   p."ciudadId" AS ciudad_id,
                   c.nombre AS ciudad,
                   p."tipoAsignacion" AS tipo_asignacion,
                   p.estado,
                   p.confirmado,
                   p.notas,
                   p."motivoCambio" AS motivo_cambio
            FROM planificacion p
            JOIN taller t ON t.id = p."tallerId"
            LEFT JOIN empresa e ON e.id = p."empresaId"
            LEFT JOIN ciudad c ON c.id = p."ciudadId"
            WHERE p.trimestre = :trimestre
            ORDER BY p.semana,
                     CASE p.dia WHEN 'L' THEN 1 WHEN 'M' THEN 2 WHEN 'X' THEN 3 WHEN 'J' THEN 4 WHEN 'V' THEN 5 END,
                     p.horario
        """),
        {"trimestre": trimestre},
    )
    rows = [dict(r) for r in result.mappings().all()]

    # Compute summary stats
    # V17: dropped OK state on planificacion. CONFIRMADO is the terminal state.
    asignados = sum(1 for r in rows if r["estado"] != "VACANTE")
    vacantes = sum(1 for r in rows if r["estado"] == "VACANTE")
    confirmados = sum(1 for r in rows if r["estado"] == "CONFIRMADO")
    cancelados = sum(1 for r in rows if r["estado"] == "CANCELADO")

    return {
        "trimestre": trimestre,
        "total_slots": len(rows),
        "asignados": asignados,
        "vacantes": vacantes,
        "confirmados": confirmados,
        "cancelados": cancelados,
        "slots": rows,
    }


@router.post("/{trimestre}/validar-asignacion", response_model=ValidarAsignacionResult)
async def validar_asignacion(
    trimestre: str,
    body: ValidarAsignacionInput,
    db: AsyncSession = Depends(get_db),
):
    """
    Validates if assigning a company to a slot violates any restrictions.
    Returns warnings but does NOT block the assignment.
    Used by the frontend to show warnings before confirming.
    """
    warnings = []
    restricciones_violadas = []

    # 1. Get slot info
    slot_result = await db.execute(
        text("""
            SELECT p.id, p.semana, p.dia, p."tallerId",
                   t.nombre AS taller_nombre, t.programa,
                   t.turno AS taller_turno,
                   COALESCE(p.horario, t.horario) AS horario
            FROM planificacion p
            JOIN taller t ON t.id = p."tallerId"
            WHERE p.id = :slot_id AND p.trimestre = :tri
        """),
        {"slot_id": body.slot_id, "tri": trimestre},
    )
    slot = slot_result.mappings().first()
    if not slot:
        raise HTTPException(404, "Slot no encontrado")

    slot = dict(slot)

    # 2. Get company info
    emp_result = await db.execute(
        text('SELECT id, nombre, "turnoPreferido" FROM empresa WHERE id = :id'),
        {"id": body.empresa_id},
    )
    empresa = emp_result.mappings().first()
    if not empresa:
        raise HTTPException(404, "Empresa no encontrada")
    empresa = dict(empresa)

    # 3. Get restrictions for this company
    rest_result = await db.execute(
        text('SELECT tipo, clave, valor, "tallerId" FROM restriccion WHERE "empresaId" = :eid'),
        {"eid": body.empresa_id},
    )
    restricciones = [dict(r) for r in rest_result.mappings().all()]

    # 4. Check solo_dia
    for r in restricciones:
        if r["clave"] == "solo_dia":
            if slot["dia"] != r["valor"]:
                dia_nombres = {"L": "Lunes", "M": "Martes", "X": "Miércoles", "J": "Jueves", "V": "Viernes"}
                restricciones_violadas.append(
                    f"solo_dia: {empresa['nombre']} solo puede {dia_nombres.get(r['valor'], r['valor'])}, "
                    f"pero el slot es {dia_nombres.get(slot['dia'], slot['dia'])}"
                )

    # 5. Check solo_taller — prioriza tallerId (FK), fallback a fuzzy por nombre
    for r in restricciones:
        if r["clave"] == "solo_taller":
            if r.get("tallerId") is not None:
                if r["tallerId"] != slot["tallerId"]:
                    restricciones_violadas.append(
                        f"solo_taller: {empresa['nombre']} solo imparte un taller específico, "
                        f"pero el taller de este slot no coincide"
                    )
            else:
                taller_nombre = slot["taller_nombre"].strip().lower()
                restriccion_valor = r["valor"].strip().lower()
                if restriccion_valor not in taller_nombre and taller_nombre not in restriccion_valor:
                    restricciones_violadas.append(
                        f"solo_taller: {empresa['nombre']} solo imparte '{r['valor']}', "
                        f"pero el taller es '{slot['taller_nombre']}'"
                    )

    # 5b. V16 — franja_horaria / franja_por_dia
    # franja_por_dia for the slot's day takes priority over franja_horaria.
    # HARD violations go to restricciones_violadas; SOFT to warnings.
    slot_horario = slot.get("horario") or ""

    # V16.1: HARD day-exclusivity from franja_por_dia. If the empresa has any
    # HARD franja_por_dia row, the declared days are the only valid ones.
    # This is reported BEFORE the franja-match check; the franja-match block
    # below only emits when the franja itself is wrong, not when the day is.
    dias_permitidos_hard = _dias_exclusivos_hard(restricciones)
    dia_fuera_set = (
        dias_permitidos_hard is not None
        and slot["dia"] not in dias_permitidos_hard
    )
    if dia_fuera_set:
        restricciones_violadas.append(
            f"franja_por_dia_dia: {empresa['nombre']} tiene franja_por_dia HARD "
            f"limitada a {sorted(dias_permitidos_hard)}, pero se asigna en {slot['dia']}"
        )

    franja_pref, tipo_franja = _franja_preferida(restricciones, slot["dia"])
    if (
        franja_pref is not None
        and slot_horario != franja_pref
        and not dia_fuera_set  # avoid double-reporting when the day itself is invalid
    ):
        msg = (
            f"franja: {empresa['nombre']} requiere franja '{franja_pref}' "
            f"el {slot['dia']}, pero el slot es de '{slot_horario}'"
        )
        if tipo_franja == "HARD":
            restricciones_violadas.append(msg)
        else:
            warnings.append(msg)

    # 6. Check no_comodin (warn that this company shouldn't be used as substitute)
    for r in restricciones:
        if r["clave"] == "no_comodin":
            restricciones_violadas.append(
                f"no_comodin: {empresa['nombre']} no debería usarse como comodín/sustituta"
            )

    # 7. Check max_extras — count how many times this company already appears beyond its frequency
    for r in restricciones:
        if r["clave"] == "max_extras":
            max_extras = int(r["valor"])
            # Count current assignments for this company
            count_result = await db.execute(
                text("""
                    SELECT COUNT(*) AS total FROM planificacion
                    WHERE trimestre = :tri AND "empresaId" = :eid
                    AND estado NOT IN ('CANCELADO', 'VACANTE')
                """),
                {"tri": trimestre, "eid": body.empresa_id},
            )
            current = count_result.scalar() or 0
            # Get original frequency
            freq_result = await db.execute(
                text("""
                    SELECT "totalAsignado" FROM frecuencia
                    WHERE trimestre = :tri AND "empresaId" = :eid
                """),
                {"tri": trimestre, "eid": body.empresa_id},
            )
            freq_row = freq_result.mappings().first()
            original_freq = int(freq_row["totalAsignado"]) if freq_row else 0
            extras_used = max(0, current - original_freq)
            if extras_used >= max_extras:
                restricciones_violadas.append(
                    f"max_extras: {empresa['nombre']} ya tiene {extras_used} extras "
                    f"(máximo permitido: {max_extras})"
                )

    # 8. Check if company already has a slot this week (H6 violation)
    week_result = await db.execute(
        text("""
            SELECT COUNT(*) AS total FROM planificacion
            WHERE trimestre = :tri AND semana = :sem AND "empresaId" = :eid
            AND estado NOT IN ('CANCELADO', 'VACANTE')
            AND id != :slot_id
        """),
        {"tri": trimestre, "sem": slot["semana"], "eid": body.empresa_id, "slot_id": body.slot_id},
    )
    week_count = week_result.scalar() or 0
    if week_count > 0:
        warnings.append(
            f"{empresa['nombre']} ya tiene {week_count} taller(es) en la semana {slot['semana']}"
        )

    # 9. Check turno preference (soft warning)
    turno_pref = empresa.get("turnoPreferido")
    if turno_pref and slot.get("taller_turno") and slot["taller_turno"] != turno_pref:
        turno_nombres = {"M": "Mañana", "T": "Tarde"}
        warnings.append(
            f"{empresa['nombre']} prefiere turno de {turno_nombres.get(turno_pref, turno_pref)}, "
            f"pero el slot es de {turno_nombres.get(slot['taller_turno'], slot['taller_turno'])}"
        )

    # Combine
    all_warnings = restricciones_violadas + warnings

    return {
        "ok": len(all_warnings) == 0,
        "warnings": all_warnings,
        "restricciones_violadas": restricciones_violadas,
    }


@router.post("/{trimestre}/exportar-excel")
async def exportar_excel(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Genera Excel del calendario con TODOS los slots (asignados + vacantes).
    Lee estado, confirmado y notas directamente de la tabla planificacion.
    """
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── 1. Cargar TODOS los slots de planificacion ────────────
    result = await db.execute(
        text("""
            SELECT p.id, p.semana, p.dia, p.horario,
                   COALESCE(p.turno, t.turno) AS turno,
                   e.nombre AS empresa,
                   e_orig.nombre AS empresa_original,
                   t.nombre AS taller,
                   t.programa,
                   c.nombre AS ciudad,
                   p."tipoAsignacion" AS tipo,
                   p.estado,
                   p.confirmado,
                   p.notas,
                   p."motivoCambio" AS motivo_cambio
            FROM planificacion p
            JOIN taller t ON t.id = p."tallerId"
            LEFT JOIN empresa e ON e.id = p."empresaId"
            LEFT JOIN empresa e_orig ON e_orig.id = p."empresaIdOriginal"
            LEFT JOIN ciudad c ON c.id = p."ciudadId"
            WHERE p.trimestre = :trimestre
            ORDER BY p.semana,
                     CASE p.dia WHEN 'L' THEN 1 WHEN 'M' THEN 2 WHEN 'X' THEN 3 WHEN 'J' THEN 4 WHEN 'V' THEN 5 END,
                     p.horario
        """),
        {"trimestre": trimestre},
    )
    db_rows = [dict(r) for r in result.mappings().all()]

    if not db_rows:
        raise HTTPException(status_code=404, detail=f"No hay slots para el trimestre {trimestre}")

    # ── 2. Construir filas para Excel ─────────────────────────
    all_rows: list[dict] = []
    has_notas = False
    has_motivo = False
    has_empresa_original = False

    for row in db_rows:
        estado = row["estado"] or "PLANIFICADO"
        confirmado = row["confirmado"]
        notas = row["notas"]
        motivo = row.get("motivo_cambio")
        empresa_original = row.get("empresa_original")

        if notas:
            has_notas = True
        if motivo:
            has_motivo = True
        # Show empresa original column if any slot has a different original empresa
        if empresa_original and empresa_original != row.get("empresa"):
            has_empresa_original = True

        # Format motivo_cambio for display
        motivo_display = ""
        if motivo == "EMPRESA_CANCELO":
            motivo_display = "Empresa canceló"
        elif motivo == "DECISION_PLANIFICADOR":
            motivo_display = "Decisión planificador"

        all_rows.append({
            "semana": row["semana"],
            "fecha": calcular_fecha_slot(trimestre, row["semana"], row["dia"]),
            "dia": row["dia"],
            "horario": row["horario"],
            "turno": row["turno"] or "",
            "empresa": row["empresa"] or "",
            "empresa_original": empresa_original or "",
            "taller": row["taller"],
            "programa": row["programa"],
            "ciudad": row["ciudad"] or "",
            "tipo": row["tipo"] or "BASE",
            "estado": estado,
            "confirmado": "SÍ" if confirmado else "",
            "notas": notas or "",
            "motivo_cambio": motivo_display,
        })

    # ── 3. Contar estados para resumen ────────────────────────
    # V17: dropped OK on planificacion. CONFIRMADO is the terminal state.
    estado_counts = {"CONFIRMADO": 0, "PLANIFICADO": 0, "CANCELADO": 0, "VACANTE": 0}
    for r in all_rows:
        estado = r["estado"]
        if estado in estado_counts:
            estado_counts[estado] += 1
        else:
            estado_counts["PLANIFICADO"] += 1

    # ── 4. Crear Excel con formato profesional ────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Calendario {trimestre}"

    # Estilos
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D3748")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    normal_font = Font(name="Arial", size=9)

    # Estado-specific styles
    # V17: OK collapsed into CONFIRMADO. CONFIRMADO inherits the former OK
    # green palette (it is now the terminal state).
    vacante_fill = PatternFill("solid", fgColor="FEF3C7")  # amarillo suave
    vacante_font = Font(name="Arial", size=9, color="92400E", italic=True)
    cancelado_fill = PatternFill("solid", fgColor="FEE2E2")  # rojo suave
    cancelado_font = Font(name="Arial", size=9, color="991B1B", strike=True)
    confirmado_fill = PatternFill("solid", fgColor="D1FAE5")  # verde suave
    confirmado_font = Font(name="Arial", size=9, color="065F46")

    # Headers
    headers = [
        "Semana", "Fecha", "Día", "Horario", "Turno", "Empresa",
    ]
    col_widths = [8, 14, 6, 14, 10, 25]

    if has_empresa_original:
        headers.append("Empresa Original")
        col_widths.append(25)

    headers.extend(["Taller", "Programa", "Ciudad", "Tipo", "Estado", "Confirmado"])
    col_widths.extend([42, 10, 12, 12, 12, 14])

    if has_notas:
        headers.append("Notas")
        col_widths.append(30)

    if has_motivo:
        headers.append("Motivo cambio")
        col_widths.append(20)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = w

    # Freeze header
    ws.freeze_panes = "A2"
    # Auto-filter — calculate last column dynamically
    col_count = len(headers)
    last_col = chr(ord('A') + col_count - 1) if col_count <= 26 else "Z"
    ws.auto_filter.ref = f"A1:{last_col}{len(all_rows) + 1}"

    # Data rows
    for i, row in enumerate(all_rows, 2):
        estado = row["estado"]
        is_vacante = estado == "VACANTE"
        is_cancelado = estado == "CANCELADO"
        is_confirmado = estado == "CONFIRMADO"

        values = [
            row["semana"], row["fecha"], row["dia"], row["horario"], row["turno"],
            row["empresa"],
        ]
        if has_empresa_original:
            values.append(row["empresa_original"])
        values.extend([
            row["taller"], row["programa"],
            row["ciudad"], row["tipo"], row["estado"], row["confirmado"],
        ])
        if has_notas:
            values.append(row["notas"])
        if has_motivo:
            values.append(row["motivo_cambio"])

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Apply estado-specific styling
            if is_vacante:
                cell.fill = vacante_fill
                cell.font = vacante_font
            elif is_cancelado:
                cell.fill = cancelado_fill
                # Only strikethrough on empresa column (6 — after adding Fecha)
                if col == 6:
                    cell.font = cancelado_font
                else:
                    cell.font = Font(name="Arial", size=9, color="991B1B")
            elif is_confirmado:
                cell.fill = confirmado_fill
                cell.font = confirmado_font
            else:
                cell.font = normal_font

            # Semana column: center and bold (CONFIRMADO uses former OK green)
            if col == 1:
                cell.font = Font(
                    name="Arial", size=9, bold=True,
                    color="92400E" if is_vacante else "991B1B" if is_cancelado else "065F46" if is_confirmado else "000000",
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── 5. Hoja resumen ───────────────────────────────────────
    ws_resumen = wb.create_sheet("Resumen")
    ws_resumen["A1"] = f"Calendario {trimestre}"
    ws_resumen["A1"].font = Font(name="Arial", bold=True, size=14)

    total_slots = len(all_rows)
    resumen_data = [
        ("Total slots", total_slots),
        ("", ""),
        ("ESTADOS", ""),
        ("Confirmados", estado_counts["CONFIRMADO"]),
        ("Planificados", estado_counts["PLANIFICADO"]),
        ("Cancelados", estado_counts["CANCELADO"]),
        ("Vacantes", estado_counts["VACANTE"]),
        ("", ""),
        ("INSTRUCCIONES", ""),
        ("1. Columna 'Empresa'", "Completar vacantes con empresa asignada"),
        ("2. Columna 'Estado'", "CONFIRMADO = confirmado con la empresa, CANCELADO = no realizado"),
        ("3. Columna 'Confirmado'", "SÍ = empresa confirmó asistencia"),
        ("4. Al cierre del trimestre", "Importar como histórico en el sistema"),
    ]

    for i, (label, value) in enumerate(resumen_data, 3):
        cell_a = ws_resumen.cell(row=i, column=1, value=label)
        cell_b = ws_resumen.cell(row=i, column=2, value=value)
        is_header = label in ("ESTADOS", "INSTRUCCIONES")
        cell_a.font = Font(name="Arial", size=10, bold=is_header or bool(label))
        cell_b.font = Font(name="Arial", size=10)

    ws_resumen.column_dimensions["A"].width = 30
    ws_resumen.column_dimensions["B"].width = 40

    # ── 6. Devolver ───────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=calendario_{trimestre}.xlsx"
        },
    )


# ── Slot Operations (Fase 3 — Operación) ────────────────────


@router.patch("/{trimestre}/slots/{slot_id}")
async def actualizar_slot(
    trimestre: str,
    slot_id: int,
    body: SlotUpdateInput,
    db: AsyncSession = Depends(get_db),
):
    """
    Update a slot's estado, confirmado, empresaId, or notas.
    Used for: confirming, cancelling, assigning company to vacancy, adding notes.
    """
    # Verify slot exists and belongs to this trimestre
    check = await db.execute(
        text('SELECT id, "empresaId", "empresaIdOriginal", estado, semana FROM planificacion WHERE id = :id AND trimestre = :tri'),
        {"id": slot_id, "tri": trimestre},
    )
    existing = check.mappings().first()
    if not existing:
        raise HTTPException(status_code=404, detail=f"Slot {slot_id} not found in {trimestre}")

    # Build dynamic UPDATE
    updates = []
    params = {"id": slot_id}

    # Handle empresa_id changes
    new_empresa_id = body.empresa_id
    # Skip H6 validation when planner explicitly sets a motivo_cambio (either reason = conscious override)
    is_forced_assignment = body.motivo_cambio in ("DECISION_PLANIFICADOR", "EMPRESA_CANCELO")

    if body.empresa_id is not None or (body.estado and body.estado == "VACANTE"):
        # If clearing empresa (setting to vacancy)
        if body.empresa_id is None and body.estado == "VACANTE":
            updates.append('"empresaId" = NULL')
        elif body.empresa_id is not None:
            # Verify company doesn't already have a slot this week (H6 constraint)
            # SKIP validation if motivo_cambio is set (planner consciously overriding)
            if not is_forced_assignment:
                week = existing["semana"]
                conflict = await db.execute(
                    text('''
                        SELECT id FROM planificacion
                        WHERE trimestre = :tri AND semana = :week AND "empresaId" = :eid AND id != :slot_id
                    '''),
                    {"tri": trimestre, "week": week, "eid": body.empresa_id, "slot_id": slot_id},
                )
                if conflict.first():
                    # Comodines (Capgemini, Indra, Santander, Repsol) pueden doblar
                    # semana en operación SOLO si se indica motivo_cambio (trazabilidad).
                    empresa_row = await db.execute(
                        text('SELECT "esComodin" FROM empresa WHERE id = :id'),
                        {"id": body.empresa_id},
                    )
                    es_comodin = bool(empresa_row.scalar() or False)
                    if es_comodin:
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"La empresa es comodín y puede doblar semana, "
                                f"pero debe indicar motivo_cambio para trazabilidad"
                            ),
                        )
                    raise HTTPException(
                        status_code=400,
                        detail=f"Company {body.empresa_id} already has a slot in week {week}",
                    )
            updates.append('"empresaId" = :empresa_id')
            params["empresa_id"] = body.empresa_id
            # WRITE-ONCE: Set empresaIdOriginal only if it's currently NULL
            # This preserves the original solver assignment for traceability
            if existing["empresaIdOriginal"] is None:
                updates.append('"empresaIdOriginal" = :empresa_id_original')
                params["empresa_id_original"] = body.empresa_id

    # Handle estado changes with auto-adjustments
    if body.estado is not None:
        new_estado = body.estado
        # Auto-adjust: if assigning empresa to a VACANTE, change estado to PLANIFICADO
        if body.empresa_id is not None and existing["estado"] == "VACANTE":
            new_estado = "PLANIFICADO"
        # Auto-adjust: if clearing empresa, change estado to VACANTE
        if body.empresa_id is None and body.estado == "VACANTE":
            new_estado = "VACANTE"
        updates.append("estado = :estado")
        params["estado"] = new_estado
    elif body.empresa_id is not None and existing["estado"] == "VACANTE":
        # Auto-transition from VACANTE to PLANIFICADO when assigning empresa
        updates.append("estado = :estado")
        params["estado"] = "PLANIFICADO"

    if body.confirmado is not None:
        updates.append("confirmado = :confirmado")
        params["confirmado"] = body.confirmado

    if body.notas is not None:
        updates.append("notas = :notas")
        params["notas"] = body.notas

    # Handle motivo_cambio
    if body.motivo_cambio is not None:
        updates.append('"motivoCambio" = :motivo_cambio')
        params["motivo_cambio"] = body.motivo_cambio
    # Auto-set motivo if estado is CANCELADO and no motivo provided (default: empresa cancelled)
    elif body.estado == "CANCELADO" and body.motivo_cambio is None:
        updates.append('"motivoCambio" = :motivo_cambio')
        params["motivo_cambio"] = "EMPRESA_CANCELO"

    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")

    updates.append('"updatedAt" = NOW()')
    query = f"UPDATE planificacion SET {', '.join(updates)} WHERE id = :id"
    await db.execute(text(query), params)
    await db.commit()

    # Return updated slot
    result = await db.execute(
        text("""
            SELECT p.id, p.semana, p.dia, p.horario,
                   COALESCE(p.turno, t.turno) AS turno,
                   p."empresaId" AS empresa_id,
                   p."empresaIdOriginal" AS empresa_id_original,
                   e.nombre AS empresa_nombre,
                   t.programa,
                   p."tallerId" AS taller_id,
                   t.nombre AS taller_nombre,
                   p."ciudadId" AS ciudad_id,
                   c.nombre AS ciudad,
                   p."tipoAsignacion" AS tipo_asignacion,
                   p.estado,
                   p.confirmado,
                   p.notas,
                   p."motivoCambio" AS motivo_cambio
            FROM planificacion p
            JOIN taller t ON t.id = p."tallerId"
            LEFT JOIN empresa e ON e.id = p."empresaId"
            LEFT JOIN ciudad c ON c.id = p."ciudadId"
            WHERE p.id = :id
        """),
        {"id": slot_id},
    )
    row = result.mappings().first()
    return {"slot": dict(row) if row else None}


@router.patch("/{trimestre}/slots-batch")
async def actualizar_slots_batch(
    trimestre: str,
    body: SlotBatchUpdateInput,
    db: AsyncSession = Depends(get_db),
):
    """
    Batch update for confirming/cancelling multiple slots at once.
    Returns: { updated: int, errors: list[str] }
    """
    updated = 0
    errors = []

    for item in body.updates:
        try:
            # Build dynamic UPDATE for each slot - include empresaIdOriginal for write-once logic
            check = await db.execute(
                text('SELECT id, "empresaIdOriginal", semana FROM planificacion WHERE id = :id AND trimestre = :tri'),
                {"id": item.slot_id, "tri": trimestre},
            )
            existing = check.mappings().first()
            if not existing:
                errors.append(f"Slot {item.slot_id} not found")
                continue

            updates = []
            params = {"id": item.slot_id}

            if item.estado is not None:
                updates.append("estado = :estado")
                params["estado"] = item.estado

            if item.confirmado is not None:
                updates.append("confirmado = :confirmado")
                params["confirmado"] = item.confirmado

            if item.empresa_id is not None:
                # H6 check: company can't have 2 slots in same week — same logic as actualizar_slot.
                # Skip if motivo_cambio was provided (planner is consciously forcing).
                # Comodines pueden doblar semana SOLO con motivo_cambio (trazabilidad).
                is_forced_assignment = item.motivo_cambio in ("DECISION_PLANIFICADOR", "EMPRESA_CANCELO")
                if not is_forced_assignment:
                    week = existing["semana"]
                    conflict = await db.execute(
                        text('''
                            SELECT id FROM planificacion
                            WHERE trimestre = :tri AND semana = :week AND "empresaId" = :eid AND id != :slot_id
                        '''),
                        {"tri": trimestre, "week": week, "eid": item.empresa_id, "slot_id": item.slot_id},
                    )
                    if conflict.first():
                        empresa_row = await db.execute(
                            text('SELECT "esComodin" FROM empresa WHERE id = :id'),
                            {"id": item.empresa_id},
                        )
                        es_comodin = bool(empresa_row.scalar() or False)
                        if es_comodin:
                            errors.append(
                                f"Slot {item.slot_id}: La empresa es comodín y puede doblar semana, "
                                f"pero debe indicar motivo_cambio para trazabilidad"
                            )
                        else:
                            errors.append(
                                f"Slot {item.slot_id}: Company {item.empresa_id} already has a slot in week {week}"
                            )
                        continue
                updates.append('"empresaId" = :empresa_id')
                params["empresa_id"] = item.empresa_id
                # WRITE-ONCE: Set empresaIdOriginal only if it's currently NULL
                if existing["empresaIdOriginal"] is None:
                    updates.append('"empresaIdOriginal" = :empresa_id_original')
                    params["empresa_id_original"] = item.empresa_id

            if item.notas is not None:
                updates.append("notas = :notas")
                params["notas"] = item.notas

            # Handle motivo_cambio
            if item.motivo_cambio is not None:
                updates.append('"motivoCambio" = :motivo_cambio')
                params["motivo_cambio"] = item.motivo_cambio
            # Auto-set motivo if estado is CANCELADO and no motivo provided
            elif item.estado == "CANCELADO" and item.motivo_cambio is None:
                updates.append('"motivoCambio" = :motivo_cambio')
                params["motivo_cambio"] = "EMPRESA_CANCELO"

            if updates:
                updates.append('"updatedAt" = NOW()')
                query = f"UPDATE planificacion SET {', '.join(updates)} WHERE id = :id"
                await db.execute(text(query), params)
                updated += 1

        except Exception as e:
            errors.append(f"Slot {item.slot_id}: {str(e)}")

    await db.commit()
    return {"updated": updated, "errors": errors}


@router.get("/{trimestre}/resumen")
async def resumen_operacion(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Returns operational summary for the quarter.
    """
    # Overall stats
    # V17: estado=OK was collapsed into CONFIRMADO. Confirmados is now the
    # terminal "completed" count for progress calculation.
    result = await db.execute(
        text("""
            SELECT
                COUNT(*) AS total_slots,
                SUM(CASE WHEN estado != 'VACANTE' THEN 1 ELSE 0 END) AS asignados,
                SUM(CASE WHEN estado = 'VACANTE' THEN 1 ELSE 0 END) AS vacantes,
                SUM(CASE WHEN estado = 'CONFIRMADO' THEN 1 ELSE 0 END) AS confirmados,
                SUM(CASE WHEN estado = 'CANCELADO' THEN 1 ELSE 0 END) AS cancelados
            FROM planificacion
            WHERE trimestre = :trimestre
        """),
        {"trimestre": trimestre},
    )
    totals = dict(result.mappings().first() or {})

    # By week
    by_week_result = await db.execute(
        text("""
            SELECT
                semana,
                COUNT(*) AS total,
                SUM(CASE WHEN estado != 'VACANTE' THEN 1 ELSE 0 END) AS asignados,
                SUM(CASE WHEN estado = 'VACANTE' THEN 1 ELSE 0 END) AS vacantes,
                SUM(CASE WHEN estado = 'CONFIRMADO' THEN 1 ELSE 0 END) AS confirmados,
                SUM(CASE WHEN estado = 'CANCELADO' THEN 1 ELSE 0 END) AS cancelados
            FROM planificacion
            WHERE trimestre = :trimestre
            GROUP BY semana
            ORDER BY semana
        """),
        {"trimestre": trimestre},
    )
    by_week = [dict(r) for r in by_week_result.mappings().all()]

    # By company
    by_company_result = await db.execute(
        text("""
            SELECT
                e.nombre AS empresa,
                COUNT(*) AS total,
                SUM(CASE WHEN p.estado = 'CONFIRMADO' THEN 1 ELSE 0 END) AS confirmados,
                SUM(CASE WHEN p.estado = 'CANCELADO' THEN 1 ELSE 0 END) AS cancelados
            FROM planificacion p
            JOIN empresa e ON e.id = p."empresaId"
            WHERE p.trimestre = :trimestre AND p."empresaId" IS NOT NULL
            GROUP BY e.nombre
            ORDER BY total DESC
        """),
        {"trimestre": trimestre},
    )
    by_company = [dict(r) for r in by_company_result.mappings().all()]

    # Progress percentage: confirmados / total (V17: confirmados now includes former-OK)
    total = totals.get("total_slots") or 0
    confirmados_count = totals.get("confirmados") or 0
    progress_pct = round((confirmados_count / total) * 100, 1) if total > 0 else 0

    return {
        "trimestre": trimestre,
        **totals,
        "progress_pct": progress_pct,
        "by_week": by_week,
        "by_company": by_company,
    }


# ── Análisis: Planificado vs Realizado ──────────────────────


@router.get("/{trimestre}/analisis", response_model=AnalisisResponse)
async def analisis_planificado_vs_realizado(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Compares solver's original assignment vs final state.
    Returns per-company metrics: cumplimiento, sustituciones, extras.

    This endpoint helps identify:
    - Companies that were substituted (unreliable)
    - Companies that stepped in as substitutes (reliable/flexible)
    - The real "compliance rate" of each company
    """
    # 1. Get all slots with both current and original empresa
    result = await db.execute(
        text("""
            SELECT
                p.id,
                p.semana,
                p.dia,
                p."tallerId" AS taller_id,
                t.nombre AS taller_nombre,
                t.programa,
                p."empresaId" AS empresa_id_final,
                e_final.nombre AS empresa_final,
                p."empresaIdOriginal" AS empresa_id_original,
                e_orig.nombre AS empresa_original,
                p.estado
            FROM planificacion p
            JOIN taller t ON t.id = p."tallerId"
            LEFT JOIN empresa e_final ON e_final.id = p."empresaId"
            LEFT JOIN empresa e_orig ON e_orig.id = p."empresaIdOriginal"
            WHERE p.trimestre = :trimestre
            ORDER BY p.semana, p.dia
        """),
        {"trimestre": trimestre},
    )
    slots = [dict(r) for r in result.mappings().all()]

    if not slots:
        raise HTTPException(status_code=404, detail=f"No hay datos para {trimestre}")

    # 2. Compute per-company metrics
    empresas_stats: dict[int, dict] = {}

    for slot in slots:
        eid_orig = slot["empresa_id_original"]
        eid_final = slot["empresa_id_final"]
        estado = slot["estado"]

        # Skip vacantes (no company assigned by solver)
        if eid_orig is None and eid_final is None:
            continue

        # Track original assignments
        if eid_orig is not None:
            if eid_orig not in empresas_stats:
                empresas_stats[eid_orig] = {
                    "empresa_id": eid_orig,
                    "empresa_nombre": slot["empresa_original"],
                    "asignados_solver": 0,
                    "cumplidos": 0,       # Slot ended with this company (CONFIRMADO)
                    "sustituida": 0,      # Slot ended with DIFFERENT company
                    "cancelados": 0,      # Slot was CANCELADO
                    "pendientes": 0,      # Slot still PLANIFICADO/CONFIRMADO
                    "extras_cubiertos": 0,  # Slots where this company REPLACED another
                }
            empresas_stats[eid_orig]["asignados_solver"] += 1

            if eid_final == eid_orig and estado == "CONFIRMADO":
                empresas_stats[eid_orig]["cumplidos"] += 1
            elif eid_final != eid_orig and eid_final is not None:
                empresas_stats[eid_orig]["sustituida"] += 1
            elif estado == "CANCELADO":
                empresas_stats[eid_orig]["cancelados"] += 1
            else:
                empresas_stats[eid_orig]["pendientes"] += 1

        # Track substitute assignments (company was NOT in solver but IS in final)
        if eid_final is not None and eid_final != eid_orig:
            if eid_final not in empresas_stats:
                empresas_stats[eid_final] = {
                    "empresa_id": eid_final,
                    "empresa_nombre": slot["empresa_final"],
                    "asignados_solver": 0,
                    "cumplidos": 0,
                    "sustituida": 0,
                    "cancelados": 0,
                    "pendientes": 0,
                    "extras_cubiertos": 0,
                }
            empresas_stats[eid_final]["extras_cubiertos"] += 1

    # 3. Compute rates
    analysis = []
    for eid, stats in empresas_stats.items():
        asignados = stats["asignados_solver"]
        tasa_cumplimiento = round(
            (stats["cumplidos"] / asignados * 100) if asignados > 0 else 0, 1
        )
        tasa_sustitucion = round(
            (stats["sustituida"] / asignados * 100) if asignados > 0 else 0, 1
        )

        analysis.append({
            **stats,
            "tasa_cumplimiento": tasa_cumplimiento,
            "tasa_sustitucion": tasa_sustitucion,
            # For next quarter: if tasa_cumplimiento < 70%, suggest reduction
            "sugerencia": (
                "REDUCIR" if tasa_cumplimiento < 70 and asignados > 0
                else "MANTENER" if tasa_cumplimiento >= 90
                else "REVISAR" if asignados > 0
                else "SOLO_COMODIN"
            ),
        })

    # Sort by tasa_cumplimiento ascending (worst first)
    analysis.sort(key=lambda x: (x["tasa_cumplimiento"], -x["asignados_solver"]))

    # 4. Global summary
    total_slots = len([s for s in slots if s["empresa_id_original"] is not None])
    total_cumplidos = sum(s["cumplidos"] for s in empresas_stats.values())
    total_sustituidos = sum(s["sustituida"] for s in empresas_stats.values())
    total_cancelados = sum(s["cancelados"] for s in empresas_stats.values())
    total_pendientes = sum(s["pendientes"] for s in empresas_stats.values())

    # Slot-level changes detail (for the changes table)
    cambios = []
    for slot in slots:
        eid_orig = slot["empresa_id_original"]
        eid_final = slot["empresa_id_final"]
        if eid_orig is not None and eid_final is not None and eid_orig != eid_final:
            cambios.append({
                "semana": slot["semana"],
                "dia": slot["dia"],
                "taller": slot["taller_nombre"],
                "programa": slot["programa"],
                "empresa_original": slot["empresa_original"],
                "empresa_final": slot["empresa_final"],
            })

    return {
        "trimestre": trimestre,
        "resumen": {
            "total_slots_asignados": total_slots,
            "cumplidos_sin_cambio": total_cumplidos,
            "sustituidos": total_sustituidos,
            "cancelados": total_cancelados,
            "pendientes": total_pendientes,
            "tasa_cumplimiento_global": round(
                (total_cumplidos / total_slots * 100) if total_slots > 0 else 0, 1
            ),
            "tasa_sustitucion_global": round(
                (total_sustituidos / total_slots * 100) if total_slots > 0 else 0, 1
            ),
        },
        "por_empresa": analysis,
        "cambios": cambios,
        "total_empresas": len(analysis),
    }


# ── Import Excel (re-import edited calendar) ────────────────


@router.post("/{trimestre}/importar-excel", response_model=ImportarExcelResult)
async def importar_excel(
    trimestre: str,
    dry_run: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """
    Importa un Excel editado de vuelta al calendario (planificacion).

    Matching: trimestre + semana + taller_nombre (slot fijo)
    Updates: empresa (por nombre), estado, confirmado

    El Excel debe tener las mismas columnas que el exportado:
    Semana, Día, Horario, Turno, Empresa, Taller, Programa, Ciudad, Tipo, Estado, Confirmado
    """
    from fastapi import UploadFile, File
    raise HTTPException(
        status_code=501,
        detail="Use POST /{trimestre}/importar-excel-file with file upload",
    )


@router.post("/{trimestre}/importar-excel-file", response_model=ImportarExcelResult)
async def importar_excel_file(
    trimestre: str,
    file: UploadFile = File(...),
    dry_run: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """
    Importa un Excel editado de vuelta al calendario (planificacion).

    Matching: trimestre + semana + taller_nombre (slot fijo)
    Updates: empresa (por nombre), estado, confirmado

    Parámetros:
    - file: Excel con el calendario editado
    - dry_run: Si true, solo muestra qué cambiaría sin aplicar cambios

    El Excel debe tener las mismas columnas que el exportado:
    Semana, Día, Horario, Turno, Empresa, Taller, Programa, Ciudad, Tipo, Estado, Confirmado
    """
    import openpyxl
    from io import BytesIO

    warnings: list[str] = []
    empresas_cambiadas: list[dict] = []
    cambios_detalle: list[dict] = []
    total_procesados = 0
    actualizados = 0
    sin_cambios = 0
    errores = 0

    # ── 1. Leer Excel ─────────────────────────────────────────
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer Excel: {str(e)}")

    # ── 2. Validar headers ────────────────────────────────────
    # V21 / Deuda 4: "Empresa Original" + "Tipo" entran al col_map para
    # desambiguar slots compartidos BASE+EXTRA. Ambas siguen siendo opcionales:
    # un Excel viejo sin "Empresa Original" cae al fallback por empresa actual,
    # y "Tipo" solo se usa para cross-validation (warning, nunca skip).
    # V21 / Deuda 4 follow-up: "Motivo cambio" también opcional. Si la columna
    # no existe o la celda viene vacía, NO se sobrescribe el motivoCambio en BD
    # (re-importar un Excel "limpio" no debe borrar motivos previos).
    expected_headers = ["Semana", "Fecha", "Día", "Horario", "Turno", "Empresa",
                        "Empresa Original", "Taller", "Programa", "Ciudad",
                        "Tipo", "Estado", "Confirmado", "Motivo cambio"]
    actual_headers = [cell.value for cell in ws[1]]

    # Normalize headers (strip whitespace, case-insensitive match)
    actual_normalized = [str(h).strip().lower() if h else "" for h in actual_headers]
    expected_normalized = [h.lower() for h in expected_headers]

    # Find column indices
    col_map = {}
    for i, expected in enumerate(expected_normalized):
        for j, actual in enumerate(actual_normalized):
            if expected == actual:
                col_map[expected_headers[i]] = j
                break

    required = ["Semana", "Taller", "Empresa", "Estado", "Confirmado"]
    missing = [h for h in required if h not in col_map]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Columnas requeridas no encontradas: {', '.join(missing)}. "
                   f"Headers encontrados: {actual_headers[:15]}",
        )

    # ── 3. Cargar mapa de empresas (nombre -> id) ─────────────
    emp_result = await db.execute(
        text("SELECT id, nombre FROM empresa WHERE activa = true")
    )
    empresas_db = {r["nombre"].strip().lower(): r["id"] for r in emp_result.mappings().all()}

    # ── 4. Cargar slots existentes indexados ──────────────────
    # V21 / Deuda 4: incluir tipoAsignacion para cross-validation contra col K.
    # V21 / Deuda 4 follow-up: incluir motivoCambio para detectar diferencias
    # contra col N y evitar UPDATEs ruidosos cuando el valor coincide.
    slots_result = await db.execute(
        text("""
            SELECT
                p.id, p.semana, p.dia, p.horario,
                p."empresaId" AS empresa_id,
                p."empresaIdOriginal" AS empresa_id_original,
                p."tipoAsignacion" AS tipo_asignacion,
                p."motivoCambio" AS motivo_cambio,
                e.nombre AS empresa_nombre,
                t.nombre AS taller_nombre,
                t.programa,
                p.estado,
                p.confirmado
            FROM planificacion p
            JOIN taller t ON t.id = p."tallerId"
            LEFT JOIN empresa e ON e.id = p."empresaId"
            WHERE p.trimestre = :tri
        """),
        {"tri": trimestre},
    )
    slots_db = [dict(r) for r in slots_result.mappings().all()]

    if not slots_db:
        raise HTTPException(
            status_code=404,
            detail=f"No hay slots de planificacion para {trimestre}",
        )

    # V21 / Deuda 4: index now keeps a list per (sem, taller) so shared slots
    # (BASE + EXTRA at the same key) can be disambiguated downstream by empresa
    # original / actual rather than collapsed to a single arbitrary row.
    slot_index: dict[tuple[int, str], list[dict]] = {}
    for s in slots_db:
        key = (s["semana"], s["taller_nombre"].strip().lower())
        slot_index.setdefault(key, []).append(s)

    # ── 5. Procesar filas del Excel ───────────────────────────
    # Estado normalization map (handles typos)
    # V17: "OK" no longer accepted on planificacion. Rows with estado=OK
    # surface a row-level error suggesting the user re-export.
    ESTADO_NORMALIZE = {
        "CANCELADO": "CANCELADO",
        "CONFIRMADO": "CONFIRMADO",
        "CONFRIMADO": "CONFIRMADO",  # common typo
        "CONFIRAMDO": "CONFIRMADO",  # another typo
        "PLANIFICADO": "PLANIFICADO",
        "PLANFICADO": "PLANIFICADO",  # typo
        "VACANTE": "VACANTE",
    }

    # V21 / Deuda 4 follow-up: mapping legible (col N) → enum BD. Same exact
    # strings the exporter emits; planners typing free-form variants get a
    # warning instead of silently mapping.
    MOTIVO_LEGIBLE_TO_ENUM = {
        "Empresa canceló": "EMPRESA_CANCELO",
        "Decisión planificador": "DECISION_PLANIFICADOR",
    }

    updates_to_apply: list[dict] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row or all(cell is None for cell in row):
            continue

        total_procesados += 1

        try:
            # Extract values
            semana_val = row[col_map["Semana"]]
            taller_val = row[col_map["Taller"]]
            empresa_val = row[col_map["Empresa"]]
            estado_val = row[col_map["Estado"]]
            confirmado_val = row[col_map["Confirmado"]]
            # V21 / Deuda 4: optional disambiguation columns. None == column
            # absent in this Excel (legacy export); empty string == column
            # present but blank for this row.
            empresa_original_val = (
                row[col_map["Empresa Original"]] if "Empresa Original" in col_map else None
            )
            tipo_val = row[col_map["Tipo"]] if "Tipo" in col_map else None
            # V21 / Deuda 4 follow-up: col N. Empty/missing → motivoCambio is
            # NOT touched; only set when the planner explicitly provides one of
            # the two legible values.
            motivo_val = row[col_map["Motivo cambio"]] if "Motivo cambio" in col_map else None

            # Parse semana
            try:
                semana = int(semana_val) if semana_val is not None else None
            except (ValueError, TypeError):
                warnings.append(f"Fila {row_num}: Semana inválida '{semana_val}'")
                errores += 1
                continue

            if semana is None or not taller_val:
                warnings.append(f"Fila {row_num}: Datos incompletos (semana o taller)")
                errores += 1
                continue

            # Normalize taller name
            taller_nombre = str(taller_val).strip().lower()

            # Find candidate rows for this (sem, taller). May be 0, 1, or 2+.
            key = (semana, taller_nombre)
            candidates: list[dict] = slot_index.get(key, [])

            if not candidates:
                warnings.append(f"Fila {row_num}: Slot no encontrado (S{semana}, {taller_val})")
                errores += 1
                continue

            # ── Process empresa (col F) ───────────────────────
            # Resolve up-front: needed both as the "new empresa" if changed AND
            # as a fallback disambiguation key when col G is absent / no match.
            empresa_nueva = str(empresa_val).strip() if empresa_val else ""
            empresa_nueva_id: int | None = None

            if empresa_nueva:
                empresa_nueva_lower = empresa_nueva.lower()
                if empresa_nueva_lower in empresas_db:
                    empresa_nueva_id = empresas_db[empresa_nueva_lower]
                else:
                    warnings.append(f"Fila {row_num}: Empresa '{empresa_nueva}' no encontrada")
                    # Don't skip — disambiguation may still match by original,
                    # and even if not we can still update estado/confirmado on
                    # the resolved row.

            # ── V21 / Deuda 4: resolve "Empresa Original" (col G) ────
            # has_original_col tracks whether col G existed AND had a value.
            # Empty string in present column = treated as missing (no constraint).
            empresa_original_text = (
                str(empresa_original_val).strip() if empresa_original_val else ""
            )
            empresa_id_original_excel: int | None = None
            has_original_col = bool(empresa_original_text)

            if has_original_col:
                eo_lower = empresa_original_text.lower()
                if eo_lower in empresas_db:
                    empresa_id_original_excel = empresas_db[eo_lower]
                else:
                    # Hard-skip: if planner provided col G with a value we don't
                    # recognise, refuse to guess — avoids silent mis-updates.
                    warnings.append(
                        f"Fila {row_num}: Empresa Original '{empresa_original_text}' "
                        f"no encontrada — verifica el Excel"
                    )
                    errores += 1
                    continue

            # ── V21 / Deuda 4: disambiguate to a single slot row ─────
            # Algorithm:
            #   - 1 candidate                     → that's it (with cross-check).
            #   - 2+ candidates + col G resolved  → match on empresa_id_original.
            #   - 2+ candidates + col G missing   → match on empresa_id (actual).
            #   - 2+ candidates and neither matches uniquely → warn + skip.
            slot: dict | None = None
            match_path: str = "single"

            if len(candidates) == 1:
                slot = candidates[0]
            else:
                if has_original_col and empresa_id_original_excel is not None:
                    by_orig = [
                        c for c in candidates
                        if c["empresa_id_original"] == empresa_id_original_excel
                    ]
                    if len(by_orig) == 1:
                        slot = by_orig[0]
                        match_path = "by_original"
                    elif len(by_orig) >= 2:
                        warnings.append(
                            f"Fila {row_num}: S{semana} {taller_val}: ambigüedad "
                            f"inesperada — múltiples filas matchean. Reportar este caso."
                        )
                        errores += 1
                        continue
                    # else len(by_orig) == 0 → fall through to actual fallback

                if slot is None:
                    by_actual = [
                        c for c in candidates if c["empresa_id"] == empresa_nueva_id
                    ]
                    if len(by_actual) == 1:
                        slot = by_actual[0]
                        match_path = (
                            "fallback_actual_no_orig_col"
                            if not has_original_col
                            else "fallback_actual_after_orig_miss"
                        )
                    elif len(by_actual) >= 2:
                        warnings.append(
                            f"Fila {row_num}: S{semana} {taller_val}: ambigüedad "
                            f"inesperada — múltiples filas matchean por actual. "
                            f"Reportar este caso."
                        )
                        errores += 1
                        continue
                    else:
                        # 0 by original AND 0 by actual.
                        warnings.append(
                            f"Fila {row_num}: S{semana} {taller_val}: ninguna fila "
                            f"del slot tiene empresa original '{empresa_original_text or '—'}' "
                            f"ni actual '{empresa_nueva or '—'}'. Posible Excel "
                            f"desactualizado o reasignación manual — edita en tabla."
                        )
                        errores += 1
                        continue

            assert slot is not None  # exhausted by the branches above
            logger.info(
                "legacy_update row=%d sem=%d taller=%r path=%s slot_id=%d "
                "tipo_bd=%s",
                row_num, semana, taller_val, match_path, slot["id"],
                slot.get("tipo_asignacion"),
            )

            # ── Cross-validate Tipo (col K) — informative warning, no skip.
            if tipo_val is not None:
                tipo_excel = str(tipo_val).strip().upper()
                tipo_bd = (slot.get("tipo_asignacion") or "").strip().upper()
                if tipo_excel and tipo_bd and tipo_excel != tipo_bd:
                    warnings.append(
                        f"Fila {row_num}: S{semana} {taller_val}: tipo del Excel "
                        f"({tipo_excel}) no coincide con BD ({tipo_bd}). "
                        f"Aplicado igualmente."
                    )

            # ── Process estado (with typo normalization) ──────
            # V17: reject "OK" with a clear row-level error so the planner
            # knows their Excel is stale (it must be re-exported).
            estado_nuevo = None
            if estado_val:
                estado_raw = str(estado_val).strip().upper()
                if estado_raw == "OK":
                    warnings.append(
                        f"Fila {row_num}: Estado 'OK' ya no es válido. "
                        f"Usa 'CONFIRMADO'. Re-exporta desde la app si tienes dudas."
                    )
                    errores += 1
                    continue
                estado_nuevo = ESTADO_NORMALIZE.get(estado_raw, estado_raw)
                # Check if it's valid after normalization
                if estado_nuevo not in ESTADO_NORMALIZE.values():
                    warnings.append(f"Fila {row_num}: Estado '{estado_val}' inválido")
                    estado_nuevo = None

            # ── Process confirmado ────────────────────────────
            confirmado_nuevo = None
            if confirmado_val is not None:
                confirmado_str = str(confirmado_val).strip().upper()
                if confirmado_str in ("SÍ", "SI", "TRUE", "1", "YES", "X"):
                    confirmado_nuevo = True
                elif confirmado_str in ("NO", "FALSE", "0", ""):
                    confirmado_nuevo = False

            # ── Process motivo cambio (col N) ─────────────────
            # V21 / Deuda 4 follow-up. None == column missing or cell empty →
            # don't touch motivoCambio in DB. Unknown free-form string → warn
            # and leave untouched. Recognised legible value → mapped to enum.
            motivo_nuevo: str | None = None
            if motivo_val is not None:
                motivo_text = str(motivo_val).strip()
                if motivo_text:
                    if motivo_text in MOTIVO_LEGIBLE_TO_ENUM:
                        motivo_nuevo = MOTIVO_LEGIBLE_TO_ENUM[motivo_text]
                    else:
                        warnings.append(
                            f"Fila {row_num}: S{semana} {taller_val}: motivo "
                            f"'{motivo_text}' no reconocido — se ignora. "
                            f"Valores válidos: 'Empresa canceló', "
                            f"'Decisión planificador'."
                        )

            # ── Get and normalize DB values ───────────────────
            empresa_anterior_id = slot["empresa_id"]
            empresa_anterior_nombre = slot["empresa_nombre"]
            # Normalize DB estado too (should already be uppercase but be safe)
            estado_anterior_raw = slot["estado"]
            estado_anterior = estado_anterior_raw.strip().upper() if estado_anterior_raw else "PLANIFICADO"
            confirmado_anterior = bool(slot["confirmado"])  # Ensure boolean
            motivo_anterior = slot.get("motivo_cambio")

            # ── Check if anything changed ─────────────────────
            changes = {}

            # Empresa change
            empresa_changed = False
            if empresa_nueva_id is not None and empresa_nueva_id != empresa_anterior_id:
                changes["empresa_id"] = empresa_nueva_id
                empresa_changed = True
            elif empresa_nueva == "" and empresa_anterior_id is not None:
                # Clearing empresa (making vacancy)
                changes["empresa_id"] = None
                changes["estado"] = "VACANTE"
                empresa_changed = True

            # Estado change (compare normalized values)
            if estado_nuevo is not None and estado_nuevo != estado_anterior:
                # Auto-adjust: if empresa is being assigned to a VACANTE, estado becomes PLANIFICADO
                if "empresa_id" in changes and changes["empresa_id"] is not None and estado_anterior == "VACANTE":
                    if estado_nuevo == "VACANTE":
                        estado_nuevo = "PLANIFICADO"
                changes["estado"] = estado_nuevo

            # Confirmado change (compare as booleans)
            if confirmado_nuevo is not None and confirmado_nuevo != confirmado_anterior:
                changes["confirmado"] = confirmado_nuevo

            # V21 / Deuda 4 follow-up: motivo change. Only added when col N
            # produced a recognised value AND it differs from BD.
            if motivo_nuevo is not None and motivo_nuevo != motivo_anterior:
                changes["motivo_cambio"] = motivo_nuevo
                logger.info(
                    "legacy_update row=%d slot_id=%d motivo=%s",
                    row_num, slot["id"], motivo_nuevo,
                )

            if not changes:
                sin_cambios += 1
                continue

            # Record detailed changes for each field that changed
            empresa_nombre_display = empresa_nueva or empresa_anterior_nombre or None

            if "estado" in changes:
                cambios_detalle.append({
                    "slot_id": slot["id"],
                    "semana": semana,
                    "dia": slot["dia"],
                    "taller_nombre": slot["taller_nombre"],
                    "empresa_nombre": empresa_nombre_display,
                    "campo": "estado",
                    "valor_anterior": estado_anterior,
                    "valor_nuevo": changes["estado"],
                })

            if "confirmado" in changes:
                cambios_detalle.append({
                    "slot_id": slot["id"],
                    "semana": semana,
                    "dia": slot["dia"],
                    "taller_nombre": slot["taller_nombre"],
                    "empresa_nombre": empresa_nombre_display,
                    "campo": "confirmado",
                    "valor_anterior": "SÍ" if confirmado_anterior else "NO",
                    "valor_nuevo": "SÍ" if changes["confirmado"] else "NO",
                })

            if empresa_changed:
                cambios_detalle.append({
                    "slot_id": slot["id"],
                    "semana": semana,
                    "dia": slot["dia"],
                    "taller_nombre": slot["taller_nombre"],
                    "empresa_nombre": empresa_nueva or "(vacante)",
                    "campo": "empresa",
                    "valor_anterior": empresa_anterior_nombre or "(vacante)",
                    "valor_nuevo": empresa_nueva or "(vacante)",
                })
                # Also record in empresas_cambiadas for backward compatibility
                empresas_cambiadas.append({
                    "slot_id": slot["id"],
                    "semana": semana,
                    "dia": slot["dia"],
                    "taller_nombre": slot["taller_nombre"],
                    "empresa_anterior": empresa_anterior_nombre,
                    "empresa_nueva": empresa_nueva or "(vacante)",
                })

            updates_to_apply.append({
                "slot_id": slot["id"],
                "changes": changes,
                "empresa_id_original": slot.get("empresa_id_original"),  # For write-once logic
            })
            actualizados += 1

        except Exception as e:
            warnings.append(f"Fila {row_num}: Error procesando: {str(e)}")
            errores += 1

    # ── 6. Apply changes (if not dry_run) ─────────────────────
    if not dry_run and updates_to_apply:
        for upd in updates_to_apply:
            slot_id = upd["slot_id"]
            changes = upd["changes"]
            current_empresa_id_original = upd.get("empresa_id_original")

            set_parts = []
            params = {"id": slot_id}

            if "empresa_id" in changes:
                if changes["empresa_id"] is None:
                    set_parts.append('"empresaId" = NULL')
                else:
                    set_parts.append('"empresaId" = :empresa_id')
                    params["empresa_id"] = changes["empresa_id"]
                    # WRITE-ONCE: Set empresaIdOriginal only if it's currently NULL
                    if current_empresa_id_original is None:
                        set_parts.append('"empresaIdOriginal" = :empresa_id_original')
                        params["empresa_id_original"] = changes["empresa_id"]

            if "estado" in changes:
                set_parts.append("estado = :estado")
                params["estado"] = changes["estado"]

            if "confirmado" in changes:
                set_parts.append("confirmado = :confirmado")
                params["confirmado"] = changes["confirmado"]

            # V21 / Deuda 4 follow-up: persist motivoCambio when set.
            if "motivo_cambio" in changes:
                set_parts.append('"motivoCambio" = :motivo_cambio')
                params["motivo_cambio"] = changes["motivo_cambio"]

            set_parts.append('"updatedAt" = NOW()')

            query = f"UPDATE planificacion SET {', '.join(set_parts)} WHERE id = :id"
            await db.execute(text(query), params)

        await db.commit()

    wb.close()

    return {
        "trimestre": trimestre,
        "total_procesados": total_procesados,
        "actualizados": actualizados,  # Show count even in dry_run (preview)
        "sin_cambios": sin_cambios,
        "errores": errores,
        "empresas_cambiadas": [EmpresaCambiada(**ec) for ec in empresas_cambiadas],
        "cambios_detalle": [CambioDetalle(**cd) for cd in cambios_detalle],
        "warnings": warnings[:50],  # Limitar warnings
    }


# ── V18: Bulk INSERT calendar importer ──────────────────────


_BULK_ESTADO_NORMALIZE = {
    "CANCELADO": "CANCELADO",
    "CONFIRMADO": "CONFIRMADO",
    "CONFRIMADO": "CONFIRMADO",
    "CONFIRAMDO": "CONFIRMADO",
    "PLANIFICADO": "PLANIFICADO",
    "PLANFICADO": "PLANIFICADO",
    "VACANTE": "VACANTE",
}

_BULK_TIPO_VALID = {"BASE", "EXTRA", "CONTINGENCIA"}


def _bulk_parse_bool(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("SI", "SÍ", "TRUE", "1", "YES", "X")


def _bulk_normalize_motivo(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    norm = s.upper().replace(" ", "_")
    if "CANCEL" in norm:
        return "EMPRESA_CANCELO"
    if "PLANIFICADOR" in norm:
        return "DECISION_PLANIFICADOR"
    if norm in ("EMPRESA_CANCELO", "DECISION_PLANIFICADOR"):
        return norm
    return None


@router.post("/{trimestre}/importar-excel-bulk", response_model=ImportarExcelBulkResult)
async def importar_excel_bulk(
    trimestre: str,
    file: UploadFile = File(...),
    dry_run: bool = False,
    wipe_first: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """
    V18: bulk INSERT importer for a prepared calendar Excel.

    Designed for loading legacy/real data into an empty trimestre.
    UPDATE-style adjustments still go through /importar-excel-file.

    Headers (case-insensitive, whitespace-stripped):
      Required: Semana, Fecha, Día, Horario, Turno, Empresa, Taller, Programa, Ciudad, Estado
      Optional: Empresa Original, Tipo, Confirmado, Notas, Motivo cambio

    Behavior:
      - wipe_first=True  → DELETE planificacion WHERE trimestre=:tri before inserting.
      - wipe_first=False + rows already exist → 409.
      - dry_run=True     → validate + count only (no DB writes; DELETE skipped).
    """
    import openpyxl
    from io import BytesIO

    warnings: list[str] = []
    total_procesados = 0
    insertados = 0
    vacantes = 0
    extras_insertados = 0
    empresa_no_encontrada = 0
    taller_no_encontrado = 0
    errores = 0
    extras_pending: list[dict] = []  # rows queued for EXTRA classification (post-INSERT)

    # ── 1. Read workbook ─────────────────────────────────────
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer Excel: {str(e)}")

    # ── 2. Header detection ──────────────────────────────────
    canonical_headers = [
        "Semana", "Fecha", "Día", "Horario", "Turno", "Empresa",
        "Taller", "Programa", "Ciudad", "Estado",
        "Empresa Original", "Tipo", "Confirmado", "Notas", "Motivo cambio",
    ]
    actual = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]

    col_map: dict[str, int] = {}
    for canon in canonical_headers:
        target = canon.lower()
        for j, h in enumerate(actual):
            if h == target:
                col_map[canon] = j
                break

    required = ["Semana", "Fecha", "Día", "Horario", "Turno",
                "Empresa", "Taller", "Programa", "Ciudad", "Estado"]
    missing = [h for h in required if h not in col_map]
    if missing:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail=f"Columnas requeridas no encontradas: {', '.join(missing)}. "
                   f"Headers encontrados: {[c.value for c in ws[1]][:20]}",
        )

    # ── 3. Pre-check: 409 unless wipe_first ──────────────────
    if not wipe_first:
        existing = await db.execute(
            text('SELECT COUNT(*) FROM planificacion WHERE trimestre = :tri'),
            {"tri": trimestre},
        )
        existing_count = existing.scalar() or 0
        if existing_count > 0:
            wb.close()
            raise HTTPException(
                status_code=409,
                detail=f"Ya existen {existing_count} filas para {trimestre}. "
                       f"Pasa wipe_first=true para reemplazarlas, o usa "
                       f"POST /{trimestre}/importar-excel-file para actualizaciones.",
            )

    # ── 4. Lookups ──────────────────────────────────────────
    emp_result = await db.execute(
        text("SELECT id, nombre FROM empresa WHERE activa = true")
    )
    empresas_map: dict[str, int] = {
        r["nombre"].strip().lower(): r["id"] for r in emp_result.mappings().all()
    }

    taller_result = await db.execute(
        text('SELECT id, nombre, "diaSemana", horario, programa FROM taller WHERE activo = true')
    )
    talleres_map: dict[tuple[str, str, str, str], int] = {}
    # V20: secondary index by (nombre, programa) → list of taller_ids, used as
    # fallback when the strict 4-tuple match fails so escuela-propia rows whose
    # day/horario differ from the catalog can still resolve a tallerId.
    talleres_by_name_prog: dict[tuple[str, str], list[int]] = {}
    for r in taller_result.mappings().all():
        key = (
            (r["nombre"] or "").strip().lower(),
            (r["diaSemana"] or "").strip().upper(),
            (r["horario"] or "").strip(),
            (r["programa"] or "").strip().upper(),
        )
        talleres_map[key] = r["id"]
        soft_key = (
            (r["nombre"] or "").strip().lower(),
            (r["programa"] or "").strip().upper(),
        )
        talleres_by_name_prog.setdefault(soft_key, []).append(r["id"])

    # V20: escuelaPropia per empresa for THIS trimestre (true/false). Empresas
    # without a configTrimestral row default to False.
    ep_result = await db.execute(
        text('''
            SELECT "empresaId", "escuelaPropia"
              FROM "configTrimestral"
             WHERE trimestre = :tri
        '''),
        {"tri": trimestre},
    )
    escuela_propia_map: dict[int, bool] = {
        r["empresaId"]: bool(r["escuelaPropia"]) for r in ep_result.mappings().all()
    }

    ciudad_result = await db.execute(text("SELECT id, nombre FROM ciudad"))
    ciudades_map: dict[str, int] = {
        r["nombre"].strip().lower(): r["id"] for r in ciudad_result.mappings().all()
    }

    # ── 5. Parse rows → inserts_to_apply ─────────────────────
    inserts_to_apply: list[dict] = []
    seen_empresa_404: set[str] = set()
    seen_taller_404: set[tuple] = set()
    seen_ciudad_404: set[str] = set()

    def _cell(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None for c in row):
            continue
        total_procesados += 1

        try:
            # Semana
            try:
                semana = int(row[col_map["Semana"]]) if row[col_map["Semana"]] is not None else None
            except (ValueError, TypeError):
                warnings.append(f"Fila {row_num}: Semana inválida '{row[col_map['Semana']]}'")
                errores += 1
                continue
            if semana is None or not (1 <= semana <= 13):
                warnings.append(f"Fila {row_num}: Semana fuera de rango [1..13]: {semana}")
                errores += 1
                continue

            # Día / Horario / Turno / Programa
            dia = (str(_cell(row, "Día") or "").strip().upper()) or None
            horario = (str(_cell(row, "Horario") or "").strip()) or None
            turno = (str(_cell(row, "Turno") or "").strip()) or None
            programa = (str(_cell(row, "Programa") or "").strip().upper()) or None

            if not dia or not horario or not programa:
                warnings.append(
                    f"Fila {row_num}: Faltan campos obligatorios (Día/Horario/Programa)"
                )
                errores += 1
                continue

            # Estado
            estado_raw = _cell(row, "Estado")
            if estado_raw is None or str(estado_raw).strip() == "":
                warnings.append(f"Fila {row_num}: Estado vacío")
                errores += 1
                continue
            estado_upper = str(estado_raw).strip().upper()
            if estado_upper == "OK":
                warnings.append(
                    f"Fila {row_num}: Estado 'OK' no válido (V17). Usa 'CONFIRMADO'."
                )
                errores += 1
                continue
            estado = _BULK_ESTADO_NORMALIZE.get(estado_upper)
            if not estado:
                warnings.append(f"Fila {row_num}: Estado '{estado_raw}' inválido")
                errores += 1
                continue

            # Empresa
            empresa_raw = _cell(row, "Empresa")
            empresa_str = str(empresa_raw).strip() if empresa_raw is not None else ""
            empresa_id: int | None = None
            if empresa_str:
                empresa_id = empresas_map.get(empresa_str.lower())
                if empresa_id is None:
                    if empresa_str.upper() not in seen_empresa_404:
                        warnings.append(
                            f"Fila {row_num}: Empresa '{empresa_str}' no encontrada — fila rechazada"
                        )
                        seen_empresa_404.add(empresa_str.upper())
                    # V20: empresa-not-found is now a hard rejection (was previously
                    # a soft warning that inserted with empresaId=NULL). Required so
                    # counter buckets are mutually exclusive and the math invariant
                    # total_procesados = insertados + vacantes + extras_insertados
                    #                  + empresa_no_encontrada + taller_no_encontrado
                    #                  + errores
                    # holds. Previous behavior silently dropped empresa assignments
                    # on misspelled names — surfacing as a rejection is safer.
                    empresa_no_encontrada += 1
                    continue
            else:
                # Blank empresa → vacancy. Force estado=VACANTE for consistency.
                if estado != "VACANTE":
                    warnings.append(
                        f"Fila {row_num}: Empresa vacía con estado '{estado}' → fuerzo VACANTE"
                    )
                    estado = "VACANTE"

            # Empresa Original (write-once at INSERT time)
            empresa_original_raw = _cell(row, "Empresa Original")
            empresa_id_original: int | None = empresa_id
            if empresa_original_raw is not None:
                eo_str = str(empresa_original_raw).strip()
                if eo_str and eo_str.lower() != empresa_str.lower():
                    eo_id = empresas_map.get(eo_str.lower())
                    if eo_id is not None:
                        empresa_id_original = eo_id
                    else:
                        warnings.append(
                            f"Fila {row_num}: Empresa Original '{eo_str}' no encontrada — usando empresaId"
                        )

            # Taller (strict 4-tuple match)
            taller_raw = _cell(row, "Taller")
            if not taller_raw or not str(taller_raw).strip():
                warnings.append(f"Fila {row_num}: Taller vacío")
                errores += 1
                continue
            taller_key = (
                str(taller_raw).strip().lower(), dia, horario, programa,
            )
            taller_id = talleres_map.get(taller_key)
            soft_match_used = False
            if taller_id is None:
                # V20: fallback — match by (nombre, programa) only.
                # If this softer match also fails, the row is genuinely invalid
                # catalog and we keep the legacy reject behavior.
                soft_key = (str(taller_raw).strip().lower(), programa)
                soft_candidates = talleres_by_name_prog.get(soft_key)
                if not soft_candidates:
                    if taller_key not in seen_taller_404:
                        warnings.append(
                            f"Fila {row_num}: Taller no encontrado para "
                            f"(nombre='{taller_raw}', día={dia}, horario={horario}, programa={programa})"
                        )
                        seen_taller_404.add(taller_key)
                    # V20: counter buckets are mutually exclusive — only
                    # taller_no_encontrado increments here (was previously
                    # double-counted with errores too).
                    taller_no_encontrado += 1
                    continue
                # Soft-matched: pick the first candidate. The row still has to
                # qualify as EXTRA (escuelaPropia + collision) — see post-loop
                # classification. Empresa is required for EXTRA classification:
                # vacancies cannot be EXTRA.
                if empresa_id is None:
                    if taller_key not in seen_taller_404:
                        warnings.append(
                            f"Fila {row_num}: Taller no encontrado por (día,horario) y "
                            f"sin empresa para clasificar como EXTRA — rechazado"
                        )
                        seen_taller_404.add(taller_key)
                    taller_no_encontrado += 1
                    continue
                taller_id = soft_candidates[0]
                soft_match_used = True

            # Ciudad
            ciudad_raw = _cell(row, "Ciudad")
            ciudad_id: int | None = None
            if ciudad_raw is not None:
                c_str = str(ciudad_raw).strip()
                if c_str:
                    ciudad_id = ciudades_map.get(c_str.lower())
                    if ciudad_id is None and c_str.upper() not in seen_ciudad_404:
                        warnings.append(
                            f"Fila {row_num}: Ciudad '{c_str}' no encontrada — ciudadId queda NULL"
                        )
                        seen_ciudad_404.add(c_str.upper())

            # Tipo asignación
            tipo_raw = _cell(row, "Tipo")
            tipo_asig = "BASE"
            if tipo_raw is not None:
                t_str = str(tipo_raw).strip().upper()
                if t_str == "HUECO":
                    tipo_asig = "CONTINGENCIA"
                elif t_str in _BULK_TIPO_VALID:
                    tipo_asig = t_str
            es_contingencia = (tipo_asig == "CONTINGENCIA")

            # Confirmado
            confirmado = _bulk_parse_bool(_cell(row, "Confirmado"))

            # Notas / Motivo cambio
            notas_raw = _cell(row, "Notas")
            notas_val = str(notas_raw).strip() if notas_raw is not None and str(notas_raw).strip() else None
            motivo_val = _bulk_normalize_motivo(_cell(row, "Motivo cambio"))

            ins_row = {
                "tri": trimestre,
                "sem": semana,
                "dia": dia,
                "horario": horario,
                "turno": turno,
                "eid": empresa_id,
                "eid_original": empresa_id_original,
                "tid": taller_id,
                "cid": ciudad_id,
                "tipo": tipo_asig,
                "contingencia": es_contingencia,
                "estado": estado,
                "confirmado": confirmado,
                "notas": notas_val,
                "motivo": motivo_val,
                # V20: bookkeeping for EXTRA classification (stripped before INSERT).
                "_soft_match_used": soft_match_used,
                "_row_num": row_num,
                "_taller_nombre": str(taller_raw).strip(),
                "_empresa_nombre": empresa_str,
            }
            inserts_to_apply.append(ins_row)
            # V20 hotfix: counter accounting fully deferred to the post-loop
            # classification pass — every row (strict OR soft match) must be
            # evaluated against the EXTRA rule before we know which bucket it
            # falls into (insertados / vacantes / extras_insertados / rejection).

        except Exception as e:
            warnings.append(f"Fila {row_num}: Error procesando: {str(e)}")
            errores += 1

    wb.close()

    # ── 5b. EXTRA classification pass (V20, hotfix) ──────────
    # EXTRA classification rule (V20, hotfix):
    # EXTRA = (empresa has escuelaPropia=true in this trimestre) AND
    #         (this row collides with another slot for same trimestre+semana+dia+horario
    #          belonging to a different empresa).
    #
    # This rule applies to EVERY row that successfully resolved a taller, regardless of
    # whether the catalog match was strict (nombre+día+horario+programa) or soft (nombre+programa).
    #
    # Rationale: extras are a semantic concept ("this slot was added on top of the standard
    # calendar by an escuela propia company"), not a structural one. They show up in two flavors:
    #   - Same workshop, same time, different company (most common — strict catalog match).
    #   - Workshop moved to a different day/time (less common — soft fallback match).
    # Both are EXTRA. The classification rule is the same for both.
    #
    # Both AND conditions are required. EP without collision = BASE. Collision without EP = BASE
    # (or rejected if row only matched via fallback).
    #
    # Collisions are computed against (a) other rows in the SAME bulk batch and
    # (b) rows already committed to planificacion for this trimestre. (b) is
    # only relevant when wipe_first=False (otherwise the trimestre is wiped and
    # cannot have pre-existing rows by construction of the 409 guard) — Option C.
    occupancy: dict[tuple[int, str, str], set[int]] = {}
    for ins in inserts_to_apply:
        if ins["eid"] is None:
            continue
        key = (ins["sem"], ins["dia"], ins["horario"])
        occupancy.setdefault(key, set()).add(ins["eid"])

    # Pre-existing rows (only matters when wipe_first=False).
    if not wipe_first:
        existing_slots = await db.execute(
            text('''
                SELECT semana, dia, horario, "empresaId"
                  FROM planificacion
                 WHERE trimestre = :tri
                   AND "empresaId" IS NOT NULL
            '''),
            {"tri": trimestre},
        )
        for r in existing_slots.mappings().all():
            key = (r["semana"], r["dia"], r["horario"])
            occupancy.setdefault(key, set()).add(r["empresaId"])

    final_inserts: list[dict] = []
    for ins in inserts_to_apply:
        empresa_id_row = ins["eid"]

        # Vacancies (no empresa) cannot be EXTRA — they are not soft-matched
        # either (the soft-match path requires empresa_id), so they always
        # come from a strict catalog match and go in as VACANTE.
        if empresa_id_row is None:
            final_inserts.append(ins)
            vacantes += 1
            continue

        is_ep = bool(escuela_propia_map.get(empresa_id_row, False))
        key = (ins["sem"], ins["dia"], ins["horario"])
        others = occupancy.get(key, set()) - {empresa_id_row}
        has_collision = len(others) > 0

        if is_ep and has_collision:
            ins["tipo"] = "EXTRA"
            extras_pending.append(ins)
            extras_insertados += 1
            final_inserts.append(ins)
            continue

        # Not EXTRA. Strict-matched rows fall through to BASE. Soft-matched
        # rows are rejected because they don't fit the catalog and don't
        # qualify as EXTRA either.
        if ins["_soft_match_used"]:
            reason = []
            if not is_ep:
                reason.append("empresa no tiene escuelaPropia=true en este trimestre")
            if not has_collision:
                reason.append("sin colisión en (semana,día,horario)")
            warnings.append(
                f"Fila {ins['_row_num']}: Taller con día/horario fuera de catálogo "
                f"para '{ins['_taller_nombre']}' rechazado — {' y '.join(reason)}"
            )
            taller_no_encontrado += 1
        else:
            final_inserts.append(ins)
            insertados += 1

    inserts_to_apply = final_inserts

    # ── 6. Apply (skip in dry_run) ───────────────────────────
    extras_detalle: list[FilaExtraInsertada] = []
    _BK_KEYS = ("_soft_match_used", "_row_num", "_taller_nombre", "_empresa_nombre")
    if not dry_run:
        if wipe_first:
            await db.execute(
                text('DELETE FROM planificacion WHERE trimestre = :tri'),
                {"tri": trimestre},
            )
        for ins in inserts_to_apply:
            params = {k: v for k, v in ins.items() if k not in _BK_KEYS}
            result = await db.execute(
                text("""
                    INSERT INTO planificacion (
                        trimestre, semana, dia, horario, turno,
                        "empresaId", "empresaIdOriginal", "tallerId", "ciudadId",
                        "tipoAsignacion", "esContingencia", estado, confirmado,
                        notas, "motivoCambio", "updatedAt"
                    ) VALUES (
                        :tri, :sem, :dia, :horario, :turno,
                        :eid, :eid_original, :tid, :cid,
                        :tipo, :contingencia, :estado, :confirmado,
                        :notas, :motivo, NOW()
                    )
                    RETURNING id
                """),
                params,
            )
            new_id = result.scalar()
            if ins.get("tipo") == "EXTRA" and new_id is not None:
                extras_detalle.append(
                    FilaExtraInsertada(
                        planificacion_id=new_id,
                        semana=ins["sem"],
                        dia=ins["dia"],
                        horario=ins["horario"],
                        taller_nombre=ins["_taller_nombre"],
                        empresa_nombre=ins["_empresa_nombre"],
                        fila_excel=ins["_row_num"],
                    )
                )
        await db.commit()
    else:
        # Dry-run: surface the EXTRAs that *would* be inserted with id=None
        # (row not committed, no real Planificacion.id exists yet).
        for ins in inserts_to_apply:
            if ins.get("tipo") == "EXTRA":
                extras_detalle.append(
                    FilaExtraInsertada(
                        planificacion_id=None,
                        semana=ins["sem"],
                        dia=ins["dia"],
                        horario=ins["horario"],
                        taller_nombre=ins["_taller_nombre"],
                        empresa_nombre=ins["_empresa_nombre"],
                        fila_excel=ins["_row_num"],
                    )
                )

    return ImportarExcelBulkResult(
        trimestre=trimestre,
        total_procesados=total_procesados,
        insertados=insertados,
        vacantes=vacantes,
        extras_insertados=extras_insertados,
        extras_detalle=extras_detalle,
        empresa_no_encontrada=empresa_no_encontrada,
        taller_no_encontrado=taller_no_encontrado,
        errores=errores,
        warnings=warnings[:100],
        dry_run=dry_run,
        wipe_first=wipe_first,
    )


# ── V20: EXTRAS listing ──────────────────────────────────────


_EXTRAS_ESTADO_WHITELIST = {"VACANTE", "PLANIFICADO", "CONFIRMADO", "CANCELADO"}


@router.get("/{trimestre}/extras", response_model=ListaExtrasResponse)
async def listar_extras(
    trimestre: str,
    estado: list[str] | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    """V20: list every Planificacion row with tipoAsignacion='EXTRA' for the trimestre.

    Joins empresa and taller for human-readable names; orders by
    (semana, dia, horario, taller_nombre).

    Optional ?estado=... query (repeatable) filters to the given V17 estados.
    Each value must be in {VACANTE, PLANIFICADO, CONFIRMADO, CANCELADO}; an
    invalid value returns 400.
    """
    estado_filter_sql = ""
    params: dict = {"tri": trimestre}
    if estado:
        invalid = [e for e in estado if e not in _EXTRAS_ESTADO_WHITELIST]
        if invalid:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Estado(s) inválido(s): {invalid}. "
                    f"Permitidos: {sorted(_EXTRAS_ESTADO_WHITELIST)}"
                ),
            )
        estado_filter_sql = " AND p.estado = ANY(:estados)"
        params["estados"] = list(estado)

    result = await db.execute(
        text(
            '''
            SELECT
                p.id,
                p.semana,
                p.dia,
                p.horario,
                p."empresaId"  AS empresa_id,
                p.estado,
                p.confirmado,
                p.notas,
                p."motivoCambio" AS motivo_cambio,
                p."createdAt"  AS created_at,
                e.nombre       AS empresa_nombre,
                t.nombre       AS taller_nombre
              FROM planificacion p
              LEFT JOIN empresa e ON e.id = p."empresaId"
              LEFT JOIN taller  t ON t.id = p."tallerId"
             WHERE p.trimestre = :tri
               AND p."tipoAsignacion" = 'EXTRA'
            '''
            + estado_filter_sql +
            '''
             ORDER BY p.semana, p.dia, p.horario, t.nombre
            '''
        ),
        params,
    )
    rows = result.mappings().all()
    extras = [
        SlotExtraResponse(
            id=r["id"],
            semana=r["semana"],
            dia=r["dia"],
            horario=r["horario"] or "",
            taller_nombre=r["taller_nombre"] or "",
            empresa_id=r["empresa_id"],
            empresa_nombre=r["empresa_nombre"],
            estado=r["estado"],
            confirmado=bool(r["confirmado"]),
            notas=r["notas"],
            motivo_cambio=r["motivo_cambio"],
            created_at=r["created_at"],
        )
        for r in rows
    ]
    return ListaExtrasResponse(
        trimestre=trimestre,
        total=len(extras),
        extras=extras,
    )


# ── Recalcular Scores ────────────────────────────────────────


@router.post("/recalcular-scores", response_model=RecalcularScoresResult)
async def recalcular_scores(db: AsyncSession = Depends(get_db)):
    """
    Recalculate all company scores from historical data.
    Can be called manually anytime — useful after importing legacy data.
    """
    from app.routers.scores import calcular_scores_trimestre

    warnings: list[str] = []
    result = await calcular_scores_trimestre(db, "", warnings)
    await db.commit()

    return {
        "empresas_actualizadas": result["empresas_actualizadas"],
        "detalle": result["detalle"],
        "warnings": warnings,
    }


# ── Cerrar Trimestre ─────────────────────────────────────────


@router.post("/{trimestre}/cerrar", response_model=CerrarTrimestreResult)
async def cerrar_trimestre(
    trimestre: str,
    body: CerrarTrimestreInput,
    db: AsyncSession = Depends(get_db),
):
    """
    Cierra el trimestre copiando los datos de planificacion a historicoTaller.

    Logica:
    1. Lee todos los slots de planificacion con estado OK o CANCELADO
    2. Calcula la fecha real desde trimestre + semana + dia
    3. Si confirmar=True, borra historicoTaller existente para ese trimestre
       e inserta los nuevos registros
    4. Si confirmar=False, solo devuelve un preview de lo que se cerraria

    Solo se copian slots con estado CONFIRMADO o CANCELADO.
    Slots VACANTE y PLANIFICADO se ignoran (no ejecutados).
    V17: el estado terminal en planificacion es CONFIRMADO (OK fue retirado).
    historicoTaller mantiene su propio enum (OK | CANCELADO) para registrar
    si el taller se llegó a impartir.
    """
    from datetime import date, timedelta

    # ── 1. Leer slots con estado final (CONFIRMADO o CANCELADO) ──────
    result = await db.execute(
        text("""
            SELECT
                p.id,
                p."empresaId" AS empresa_id,
                p."empresaIdOriginal" AS empresa_id_original,
                p."tallerId" AS taller_id,
                p.semana,
                p.dia,
                p.estado,
                p."motivoCambio" AS motivo_cambio,
                c.nombre AS ciudad
            FROM planificacion p
            LEFT JOIN ciudad c ON c.id = p."ciudadId"
            WHERE p.trimestre = :tri
            AND p.estado IN ('CONFIRMADO', 'CANCELADO')
            AND p."empresaId" IS NOT NULL
        """),
        {"tri": trimestre},
    )
    slots_finales = [dict(r) for r in result.mappings().all()]

    # ── 2. Contar ignorados (VACANTE + PLANIFICADO) ──────────
    ignorados_result = await db.execute(
        text("""
            SELECT COUNT(*) FROM planificacion
            WHERE trimestre = :tri
            AND (estado NOT IN ('CONFIRMADO', 'CANCELADO') OR "empresaId" IS NULL)
        """),
        {"tri": trimestre},
    )
    total_ignorado = ignorados_result.scalar() or 0

    # ── 3. Separar CONFIRMADO y CANCELADO ────────────────────
    slots_ok = [s for s in slots_finales if s["estado"] == "CONFIRMADO"]
    slots_cancelado = [s for s in slots_finales if s["estado"] == "CANCELADO"]

    total_ok = len(slots_ok)
    total_cancelado = len(slots_cancelado)

    # ── 4. Si no hay nada que cerrar, advertir ───────────────
    if total_ok == 0 and total_cancelado == 0:
        raise HTTPException(
            status_code=400,
            detail=f"No hay slots con estado CONFIRMADO o CANCELADO en {trimestre}. "
                   "El trimestre parece no estar listo para cerrar.",
        )

    # ── 5. Si es preview, devolver sin modificar ─────────────
    if not body.confirmar:
        return {
            "trimestre": trimestre,
            "total_ok": total_ok,
            "total_cancelado": total_cancelado,
            "total_ignorado": total_ignorado,
            "preview": True,
        }

    # ── 6. Calcular fechas y preparar inserts ────────────────
    def calcular_fecha(trimestre: str, semana: int, dia: str) -> date:
        """
        Calcula la fecha real desde trimestre + semana + dia.
        Q1: primer lunes de Enero, Q2: primer lunes de Abril, etc.
        """
        year = int(trimestre[:4])
        quarter = int(trimestre[-1])
        month_start = {1: 1, 2: 4, 3: 7, 4: 10}[quarter]
        first_day = date(year, month_start, 1)
        # Encontrar primer lunes
        days_until_monday = (7 - first_day.weekday()) % 7
        if days_until_monday == 0 and first_day.weekday() != 0:
            days_until_monday = 7
        first_monday = first_day + timedelta(days=days_until_monday)
        # Si el primer dia del mes es lunes, usar ese
        if first_day.weekday() == 0:
            first_monday = first_day
        # Offset por semana (semana 1 = primera semana)
        week_start = first_monday + timedelta(weeks=semana - 1)
        # Offset por dia
        dia_offset = {"L": 0, "M": 1, "X": 2, "J": 3, "V": 4}
        return week_start + timedelta(days=dia_offset.get(dia, 0))

    # ── 7. Borrar historico existente para este trimestre ────
    await db.execute(
        text('DELETE FROM "historicoTaller" WHERE trimestre = :tri'),
        {"tri": trimestre},
    )

    # ── 8. Insertar nuevos registros ─────────────────────────
    # V17: planificacion.estado = CONFIRMADO maps to historicoTaller.estado = OK
    # (the EstadoTaller enum on historicoTaller stays OK | CANCELADO — separate
    # system that tracks "did the workshop happen", untouched by V17).
    for slot in slots_finales:
        fecha = calcular_fecha(trimestre, slot["semana"], slot["dia"])
        estado_db = "OK" if slot["estado"] == "CONFIRMADO" else "CANCELADO"

        await db.execute(
            text("""
                INSERT INTO "historicoTaller" (
                    "empresaId", "empresaIdOriginal", "tallerId", fecha, estado, ciudad, trimestre, "motivoCambio", "createdAt"
                )
                VALUES (:eid, :eid_original, :tid, :fecha, :estado, :ciudad, :tri, :motivo, NOW())
            """),
            {
                "eid": slot["empresa_id"],
                "eid_original": slot.get("empresa_id_original"),
                "tid": slot["taller_id"],
                "fecha": fecha,
                "estado": estado_db,
                "ciudad": slot["ciudad"] or "MADRID",
                "tri": trimestre,
                "motivo": slot.get("motivo_cambio"),
            },
        )

    await db.commit()

    # ── 8b. Auto-limpiar esNueva en empresas con histórico ──────
    # Tras el primer trimestre cerrado, una empresa "nueva" ya no lo es.
    es_nueva_result = await db.execute(
        text('''
            UPDATE empresa
               SET "esNueva" = false
             WHERE "esNueva" = true
               AND EXISTS (
                   SELECT 1 FROM "historicoTaller" h
                    WHERE h."empresaId" = empresa.id
                    LIMIT 1
               )
             RETURNING id, nombre
        '''),
    )
    limpiadas = es_nueva_result.mappings().all()
    for row in limpiadas:
        try:
            import logging
            logging.getLogger(__name__).info(
                f"[CIERRE_TRIMESTRE] {row['nombre']} ya no es nueva "
                "(aparece en históricos)"
            )
        except Exception:
            pass
    await db.commit()

    # ── 9. Auto-calculate scores based on all historical data ────
    from app.routers.scores import calcular_scores_trimestre

    score_warnings: list[str] = []
    score_result = await calcular_scores_trimestre(db, trimestre, score_warnings)
    await db.commit()

    return {
        "trimestre": trimestre,
        "total_ok": total_ok,
        "total_cancelado": total_cancelado,
        "total_ignorado": total_ignorado,
        "preview": False,
        "scores_actualizados": score_result["empresas_actualizadas"],
        "score_warnings": score_warnings,
    }
