"""
Histórico — Importar y consultar datos históricos de talleres.
Columnas en camelCase (Prisma). Tabla: "historicoTaller".
"""

from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from io import BytesIO

from app.db import get_db

router = APIRouter()


# ── New: List distinct trimestres ────────────────────────────

@router.get("/trimestres")
async def listar_trimestres(db: AsyncSession = Depends(get_db)):
    """
    Returns list of distinct trimestres that have historical data.
    Sorted descending (most recent first).
    """
    result = await db.execute(
        text("""
            SELECT DISTINCT trimestre
            FROM "historicoTaller"
            WHERE trimestre IS NOT NULL
            ORDER BY trimestre DESC
        """)
    )
    trimestres = [row[0] for row in result.fetchall()]
    return {"trimestres": trimestres}


# ── New: Get historico for specific trimestre ────────────────

@router.get("/{trimestre}")
async def obtener_historico_trimestre(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Returns all historical records for a specific trimestre.
    Includes empresa name, taller name, programa, etc.
    """
    result = await db.execute(
        text("""
            SELECT
                h.id,
                h."empresaId" AS empresa_id,
                e.nombre AS empresa_nombre,
                h."tallerId" AS taller_id,
                t.nombre AS taller_nombre,
                t.programa,
                t."diaSemana" AS dia,
                t.turno,
                t.horario,
                h.fecha,
                h.estado,
                h.ciudad,
                h.trimestre,
                EXTRACT(WEEK FROM h.fecha) AS semana_abs
            FROM "historicoTaller" h
            JOIN empresa e ON e.id = h."empresaId"
            JOIN taller t ON t.id = h."tallerId"
            WHERE h.trimestre = :trimestre
            ORDER BY h.fecha ASC, t.turno ASC
        """),
        {"trimestre": trimestre},
    )
    registros = [dict(r) for r in result.mappings().all()]

    # Calculate stats
    total = len(registros)
    ok_count = sum(1 for r in registros if r["estado"] == "OK")
    cancelados = sum(1 for r in registros if r["estado"] == "CANCELADO")

    return {
        "trimestre": trimestre,
        "total": total,
        "ok": ok_count,
        "cancelados": cancelados,
        "registros": registros,
    }


# ── New: Export to Excel ─────────────────────────────────────

@router.post("/{trimestre}/exportar-excel")
async def exportar_historico_excel(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Generates an Excel file from historicoTaller for the given trimestre.
    Professional format with colors, filters, and summary sheet.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Fetch data
    result = await db.execute(
        text("""
            SELECT
                h.id,
                e.nombre AS empresa,
                t.nombre AS taller,
                t.programa,
                t."diaSemana" AS dia,
                t.turno,
                t.horario,
                h.fecha,
                h.estado,
                h.ciudad
            FROM "historicoTaller" h
            JOIN empresa e ON e.id = h."empresaId"
            JOIN taller t ON t.id = h."tallerId"
            WHERE h.trimestre = :trimestre
            ORDER BY h.fecha ASC, t.turno ASC
        """),
        {"trimestre": trimestre},
    )
    registros = [dict(r) for r in result.mappings().all()]

    if not registros:
        # Return empty Excel with message
        wb = Workbook()
        ws = wb.active
        ws.title = "Sin datos"
        ws["A1"] = f"No hay registros históricos para {trimestre}"
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=historico_{trimestre}.xlsx"},
        )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"

    # Styles
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_font = Font(name="Arial", size=9)
    cancelado_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    cancelado_font = Font(name="Arial", size=9, strikethrough=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # Headers
    headers = ["Fecha", "Día", "Horario", "Turno", "Empresa", "Taller", "Programa", "Ciudad", "Estado"]
    col_widths = [12, 6, 12, 8, 25, 40, 10, 12, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(registros) + 1}"

    # Data rows
    for i, reg in enumerate(registros, 2):
        is_cancelado = reg["estado"] == "CANCELADO"
        fecha_str = reg["fecha"].strftime("%d/%m/%Y") if reg["fecha"] else ""

        values = [
            fecha_str,
            reg["dia"],
            reg["horario"],
            reg["turno"],
            reg["empresa"],
            reg["taller"],
            reg["programa"],
            reg["ciudad"],
            reg["estado"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if is_cancelado:
                cell.fill = cancelado_fill
                if col == 5:  # Empresa column
                    cell.font = cancelado_font
                else:
                    cell.font = normal_font
            else:
                cell.font = normal_font

    # Summary sheet
    ws_resumen = wb.create_sheet("Resumen")
    ws_resumen["A1"] = f"Histórico {trimestre}"
    ws_resumen["A1"].font = Font(name="Arial", bold=True, size=14)

    ok_count = sum(1 for r in registros if r["estado"] == "OK")
    cancelados = sum(1 for r in registros if r["estado"] == "CANCELADO")
    total = len(registros)

    resumen_data = [
        ("Total registros", total),
        ("OK", ok_count),
        ("Cancelados", cancelados),
        ("% Asistencia", f"{round(ok_count / total * 100, 1)}%" if total > 0 else "0%"),
    ]

    for i, (label, value) in enumerate(resumen_data, 3):
        ws_resumen.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
        ws_resumen.cell(row=i, column=2, value=value).font = Font(name="Arial", size=10)

    # By company
    ws_resumen.cell(row=8, column=1, value="Por empresa").font = Font(name="Arial", bold=True, size=12)
    empresas_stats = {}
    for r in registros:
        emp = r["empresa"]
        if emp not in empresas_stats:
            empresas_stats[emp] = {"ok": 0, "cancelados": 0}
        if r["estado"] == "OK":
            empresas_stats[emp]["ok"] += 1
        else:
            empresas_stats[emp]["cancelados"] += 1

    row = 9
    ws_resumen.cell(row=row, column=1, value="Empresa").font = Font(name="Arial", bold=True, size=9)
    ws_resumen.cell(row=row, column=2, value="OK").font = Font(name="Arial", bold=True, size=9)
    ws_resumen.cell(row=row, column=3, value="Canc.").font = Font(name="Arial", bold=True, size=9)
    ws_resumen.cell(row=row, column=4, value="Total").font = Font(name="Arial", bold=True, size=9)

    for emp, stats in sorted(empresas_stats.items()):
        row += 1
        ws_resumen.cell(row=row, column=1, value=emp)
        ws_resumen.cell(row=row, column=2, value=stats["ok"])
        ws_resumen.cell(row=row, column=3, value=stats["cancelados"])
        ws_resumen.cell(row=row, column=4, value=stats["ok"] + stats["cancelados"])

    ws_resumen.column_dimensions["A"].width = 30
    ws_resumen.column_dimensions["B"].width = 10
    ws_resumen.column_dimensions["C"].width = 10
    ws_resumen.column_dimensions["D"].width = 10

    # Return Excel
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=historico_{trimestre}.xlsx"},
    )


@router.get("/")
async def listar_historico(
    trimestre: str | None = None,
    empresa_id: int | None = None,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
):
    """Consulta el histórico con filtros opcionales."""
    conditions = []
    params: dict = {"limit": limit}

    if trimestre:
        conditions.append('h.trimestre = :trimestre')
        params["trimestre"] = trimestre
    if empresa_id:
        conditions.append('h."empresaId" = :empresa_id')
        params["empresa_id"] = empresa_id

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    query = text(f"""
        SELECT h.id, h."empresaId", e.nombre AS empresa_nombre,
               h."tallerId", t.nombre AS taller_nombre,
               h.fecha, h.estado, h.ciudad, h.trimestre
        FROM "historicoTaller" h
        JOIN empresa e ON e.id = h."empresaId"
        JOIN taller t ON t.id = h."tallerId"
        {where}
        ORDER BY h.fecha DESC
        LIMIT :limit
    """)
    result = await db.execute(query, params)
    return {"historico": [dict(r) for r in result.mappings().all()]}


@router.post("/importar")
async def importar_historico(
    archivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Importa histórico desde CSV/Excel.
    Columnas esperadas: Id_Empresa, Id_Taller, Fecha_Taller, Estado, Ciudad
    """
    import pandas as pd
    from io import BytesIO

    content = await archivo.read()
    filename = archivo.filename or ""

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        df = pd.read_excel(BytesIO(content))
    else:
        df = pd.read_csv(BytesIO(content))

    column_map = {
        "Id_Empresa": "empresaId",
        "Id_Taller": "tallerId",
        "Fecha_Taller": "fecha",
        "Estado": "estado",
        "Ciudad": "ciudad",
    }
    df = df.rename(columns={k: v for k, v in column_map.items() if k in df.columns})

    inserted = 0
    errors = []

    for idx, row in df.iterrows():
        try:
            await db.execute(
                text("""
                    INSERT INTO "historicoTaller"
                        ("empresaId", "tallerId", fecha, estado, ciudad)
                    VALUES (:empresaId, :tallerId, :fecha, :estado, :ciudad)
                """),
                {
                    "empresaId": int(row["empresaId"]),
                    "tallerId": int(row["tallerId"]),
                    "fecha": pd.to_datetime(row["fecha"]).date(),
                    "estado": row.get("estado", "OK"),
                    "ciudad": row.get("ciudad"),
                },
            )
            inserted += 1
        except Exception as e:
            errors.append(f"Fila {idx}: {str(e)}")

    await db.commit()
    return {
        "insertados": inserted,
        "errores": len(errors),
        "detalle_errores": errors[:10],
    }


@router.get("/stats/{trimestre}")
async def stats_trimestre(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """Estadísticas agregadas de un trimestre para Fase 1."""
    query = text("""
        SELECT e.id AS empresa_id, e.nombre,
               COUNT(*) AS total,
               SUM(CASE WHEN h.estado = 'OK' THEN 1 ELSE 0 END) AS ok,
               SUM(CASE WHEN h.estado = 'CANCELADO' THEN 1 ELSE 0 END) AS cancelados,
               ROUND(
                   SUM(CASE WHEN h.estado = 'OK' THEN 1 ELSE 0 END)::numeric
                   / NULLIF(COUNT(*), 0), 2
               ) AS tasa_cumplimiento
        FROM "historicoTaller" h
        JOIN empresa e ON e.id = h."empresaId"
        WHERE h.trimestre = :trimestre
        GROUP BY e.id, e.nombre
        ORDER BY tasa_cumplimiento ASC
    """)
    result = await db.execute(query, {"trimestre": trimestre})
    return {"stats": [dict(r) for r in result.mappings().all()]}