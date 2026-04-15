"""
CONFIG TRIMESTRAL — CRUD para configuraciones trimestrales por empresa

Endpoints:
  GET  /{trimestre}             → Lista todas las configs del trimestre
  GET  /{trimestre}/resumen     → Resumen rápido
  GET  /{trimestre}/exportar-excel → Exporta configs a Excel (formato ideal)
  POST /{trimestre}/importar-excel → Importa Excel (detecta formato ideal/legacy)
  PUT  /{trimestre}/batch       → Actualiza múltiples configs
  POST /{trimestre}/inicializar → Inicializa configs (clonar o crear default)
  PUT  /{trimestre}/{empresaId} → Actualiza config de una empresa

IMPORTANT: Route order matters! More specific routes (/batch, /resumen, /inicializar,
/exportar-excel, /importar-excel) must be defined BEFORE the catch-all /{empresaId} route.

Tablas: configTrimestral, empresa
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from pydantic import BaseModel
from typing import Optional
from io import BytesIO
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.db import get_db

router = APIRouter()


# ── Schemas ──────────────────────────────────────────────────

class ConfigTrimestralOut(BaseModel):
    id: int
    empresa_id: int
    empresa_nombre: str
    tipo_participacion: str  # EF, IT, AMBAS
    escuela_propia: bool
    frecuencia_solicitada: Optional[int]
    disponibilidad_dias: str  # "L,M,X,J,V"
    turno_preferido: Optional[str]  # "M", "T", null
    voluntarios_disponibles: int
    preferencias_taller: Optional[str]
    notas: Optional[str]


class ConfigTrimestralUpdate(BaseModel):
    tipo_participacion: Optional[str] = None
    escuela_propia: Optional[bool] = None
    frecuencia_solicitada: Optional[int] = None
    disponibilidad_dias: Optional[str] = None
    turno_preferido: Optional[str] = None
    voluntarios_disponibles: Optional[int] = None
    preferencias_taller: Optional[str] = None
    notas: Optional[str] = None


class ConfigBatchUpdateItem(BaseModel):
    empresa_id: int
    tipo_participacion: Optional[str] = None
    escuela_propia: Optional[bool] = None
    frecuencia_solicitada: Optional[int] = None
    disponibilidad_dias: Optional[str] = None
    turno_preferido: Optional[str] = None
    voluntarios_disponibles: Optional[int] = None
    preferencias_taller: Optional[str] = None
    notas: Optional[str] = None


class ConfigBatchUpdateInput(BaseModel):
    updates: list[ConfigBatchUpdateItem]


class InicializarInput(BaseModel):
    origen_trimestre: Optional[str] = None  # Si se proporciona, clona desde ese trimestre


class InicializarResult(BaseModel):
    trimestre: str
    total_configs: int
    clonadas: int
    nuevas: int
    warnings: list[str]


class ConfigResumen(BaseModel):
    trimestre: str
    total_configs: int
    por_tipo: dict[str, int]  # EF, IT, AMBAS
    con_frecuencia: int
    sin_frecuencia: int
    escuela_propia: int


class ImportPreviewItem(BaseModel):
    empresa_id: int
    nombre: str
    frecuencia: int
    tipo: Optional[str] = None
    notas: Optional[str] = None
    # Legacy-only extra fields:
    detalle_ef: Optional[int] = None
    detalle_it: Optional[int] = None


class ImportarExcelResponse(BaseModel):
    trimestre: str
    formato_detectado: str  # "ideal" | "legacy"
    total_procesados: int
    aplicados: int  # 0 if dry_run
    preview: list[ImportPreviewItem]
    warnings: list[str]
    dry_run: bool


# ── Helper Functions ────────────────────────────────────────

# Non-Madrid company suffixes to skip in legacy format
LEGACY_SKIP_SUFFIXES = ("BCN", "SEV", "VLC", "ZGZ", "BAL")


def detect_config_format(ws) -> str:
    """
    Auto-detect the Excel format.

    Returns "ideal" if row 1 headers contain "empresa" AND "frecuencia".
    Returns "legacy" if rows 1-3 contain "trim" or "escuela" or "fortalecimiento".
    Defaults to "ideal".
    """
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        rows.append([str(c).strip().lower() if c else "" for c in row])

    if len(rows) >= 1:
        header_text = " ".join(rows[0])
        if "empresa" in header_text and "frecuencia" in header_text:
            return "ideal"

    for row in rows:
        row_text = " ".join(row)
        if "trim" in row_text or "escuela" in row_text or "fortalecimiento" in row_text:
            return "legacy"

    return "ideal"


def parse_legacy_frequency(value) -> int:
    """
    Parse legacy format frequency strings like "3 + MP" → 3.
    Extracts first integer found, or 0 if none.
    """
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if not s:
        return 0
    match = re.search(r"\d+", s)
    return int(match.group()) if match else 0


def normalize_empresa_name(name: str) -> str:
    """Normalize empresa name for fuzzy matching."""
    # Remove MAD suffix for legacy format matching
    normalized = name.strip()
    if normalized.upper().endswith(" MAD"):
        normalized = normalized[:-4].strip()
    return normalized.lower()


# ── Endpoints ────────────────────────────────────────────────
# NOTE: Order matters! Specific routes must come before parameterized routes.


@router.get("/{trimestre}")
async def obtener_configs_trimestre(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Devuelve todas las configuraciones trimestrales para el trimestre dado,
    unidas con el nombre de la empresa.
    """
    result = await db.execute(
        text("""
            SELECT
                ct.id,
                ct."empresaId" AS empresa_id,
                e.nombre AS empresa_nombre,
                ct."tipoParticipacion" AS tipo_participacion,
                ct."escuelaPropia" AS escuela_propia,
                ct."frecuenciaSolicitada" AS frecuencia_solicitada,
                ct."disponibilidadDias" AS disponibilidad_dias,
                ct."turnoPreferido" AS turno_preferido,
                ct."voluntariosDisponibles" AS voluntarios_disponibles,
                ct."preferenciasTaller" AS preferencias_taller,
                ct.notas
            FROM "configTrimestral" ct
            JOIN empresa e ON e.id = ct."empresaId"
            WHERE ct.trimestre = :trimestre
            ORDER BY e.nombre
        """),
        {"trimestre": trimestre},
    )
    configs = [dict(r) for r in result.mappings().all()]

    return {
        "trimestre": trimestre,
        "total": len(configs),
        "configs": configs,
    }


@router.get("/{trimestre}/resumen")
async def resumen_configs(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Devuelve un resumen rápido de las configuraciones del trimestre.
    """
    # Total y por tipo
    result = await db.execute(
        text("""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN "tipoParticipacion" = 'EF' THEN 1 ELSE 0 END) AS ef,
                SUM(CASE WHEN "tipoParticipacion" = 'IT' THEN 1 ELSE 0 END) AS it,
                SUM(CASE WHEN "tipoParticipacion" = 'AMBAS' THEN 1 ELSE 0 END) AS ambas,
                SUM(CASE WHEN "frecuenciaSolicitada" IS NOT NULL AND "frecuenciaSolicitada" > 0 THEN 1 ELSE 0 END) AS con_freq,
                SUM(CASE WHEN "frecuenciaSolicitada" IS NULL OR "frecuenciaSolicitada" = 0 THEN 1 ELSE 0 END) AS sin_freq,
                SUM(CASE WHEN "escuelaPropia" = true THEN 1 ELSE 0 END) AS escuela_propia
            FROM "configTrimestral"
            WHERE trimestre = :tri
        """),
        {"tri": trimestre},
    )
    row = result.mappings().first()

    if not row or row["total"] == 0:
        return {
            "trimestre": trimestre,
            "total_configs": 0,
            "por_tipo": {"EF": 0, "IT": 0, "AMBAS": 0},
            "con_frecuencia": 0,
            "sin_frecuencia": 0,
            "escuela_propia": 0,
        }

    return {
        "trimestre": trimestre,
        "total_configs": row["total"] or 0,
        "por_tipo": {
            "EF": row["ef"] or 0,
            "IT": row["it"] or 0,
            "AMBAS": row["ambas"] or 0,
        },
        "con_frecuencia": row["con_freq"] or 0,
        "sin_frecuencia": row["sin_freq"] or 0,
        "escuela_propia": row["escuela_propia"] or 0,
    }


@router.get("/{trimestre}/exportar-excel")
async def exportar_excel(
    trimestre: str,
    db: AsyncSession = Depends(get_db),
):
    """
    Exporta las configuraciones trimestrales a Excel (formato ideal).
    Columns: Empresa, Frecuencia, Tipo, Dias, Turno, Voluntarios, Notas
    """
    result = await db.execute(
        text("""
            SELECT
                e.nombre AS empresa,
                ct."frecuenciaSolicitada" AS frecuencia,
                ct."tipoParticipacion" AS tipo,
                ct."disponibilidadDias" AS dias,
                ct."turnoPreferido" AS turno,
                ct."voluntariosDisponibles" AS voluntarios,
                ct.notas
            FROM "configTrimestral" ct
            JOIN empresa e ON e.id = ct."empresaId"
            WHERE ct.trimestre = :tri
            ORDER BY e.nombre
        """),
        {"tri": trimestre},
    )
    rows = [dict(r) for r in result.mappings().all()]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ConfigTrimestral"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = ["Empresa", "Frecuencia", "Tipo", "Dias", "Turno", "Voluntarios", "Notas"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row_data["empresa"])
        ws.cell(row=row_idx, column=2, value=row_data["frecuencia"] or 0)
        ws.cell(row=row_idx, column=3, value=row_data["tipo"] or "AMBAS")
        ws.cell(row=row_idx, column=4, value=row_data["dias"] or "L,M,X,J,V")
        ws.cell(row=row_idx, column=5, value=row_data["turno"] or "-")
        ws.cell(row=row_idx, column=6, value=row_data["voluntarios"] or 0)
        ws.cell(row=row_idx, column=7, value=row_data["notas"] or "")

    # Adjust column widths
    ws.column_dimensions["A"].width = 35  # Empresa
    ws.column_dimensions["B"].width = 12  # Frecuencia
    ws.column_dimensions["C"].width = 10  # Tipo
    ws.column_dimensions["D"].width = 15  # Dias
    ws.column_dimensions["E"].width = 8   # Turno
    ws.column_dimensions["F"].width = 12  # Voluntarios
    ws.column_dimensions["G"].width = 40  # Notas

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"config_trimestral_{trimestre}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/{trimestre}/importar-excel")
async def importar_excel(
    trimestre: str,
    file: UploadFile = File(...),
    dry_run: bool = Form(True),
    db: AsyncSession = Depends(get_db),
):
    """
    Importa configuraciones desde Excel.
    Auto-detecta formato:
      - "ideal": sistema propio (Empresa, Frecuencia, Tipo, Dias, Turno, Voluntarios, Notas)
      - "legacy": Excel planificador anual (multi-quarter EF/IT split)

    dry_run=True returns preview without applying. dry_run=False applies changes.
    """
    # Load the Excel file
    try:
        content = await file.read()
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {e}")

    # Detect format
    formato = detect_config_format(ws)

    # Load empresas for matching
    emp_result = await db.execute(
        text("""
            SELECT id, nombre FROM empresa WHERE activa = true
        """)
    )
    empresas = {normalize_empresa_name(r["nombre"]): {"id": r["id"], "nombre": r["nombre"]}
                for r in emp_result.mappings().all()}

    preview = []
    warnings = []
    total_procesados = 0

    if formato == "ideal":
        # ── Process Ideal Format ──────────────────────────────
        # Find header row (should be row 1)
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[str(cell.value).strip().lower()] = col_idx

        # Required column
        empresa_col = headers.get("empresa")
        if not empresa_col:
            raise HTTPException(status_code=400, detail="Columna 'Empresa' no encontrada")

        freq_col = headers.get("frecuencia")
        tipo_col = headers.get("tipo")
        dias_col = headers.get("dias")
        turno_col = headers.get("turno")
        vol_col = headers.get("voluntarios")
        notas_col = headers.get("notas")

        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[empresa_col - 1]:
                continue

            empresa_name = str(row[empresa_col - 1]).strip()
            normalized = normalize_empresa_name(empresa_name)
            total_procesados += 1

            # Find empresa
            match = empresas.get(normalized)
            if not match:
                # Try fuzzy: check if any empresa contains this name or vice versa
                for key, val in empresas.items():
                    if normalized in key or key in normalized:
                        match = val
                        break
            if not match:
                warnings.append(f"Fila {row_idx}: Empresa '{empresa_name}' no encontrada")
                continue

            # Parse values
            frecuencia = parse_legacy_frequency(row[freq_col - 1] if freq_col else None)
            tipo = str(row[tipo_col - 1]).strip().upper() if tipo_col and row[tipo_col - 1] else None
            dias = str(row[dias_col - 1]).strip() if dias_col and row[dias_col - 1] else None
            turno = str(row[turno_col - 1]).strip() if turno_col and row[turno_col - 1] else None
            voluntarios = int(row[vol_col - 1]) if vol_col and row[vol_col - 1] else None
            notas = str(row[notas_col - 1]).strip() if notas_col and row[notas_col - 1] else None

            # Validate tipo
            if tipo and tipo not in ("EF", "IT", "AMBAS"):
                tipo = None

            # Validate turno
            if turno and turno not in ("M", "T"):
                turno = None

            preview.append(ImportPreviewItem(
                empresa_id=match["id"],
                nombre=match["nombre"],
                frecuencia=frecuencia,
                tipo=tipo,
                notas=notas,
            ))

    else:
        # ── Process Legacy Format ─────────────────────────────
        # Parse quarter number from trimestre (e.g., "2026-Q2" → 2)
        quarter_match = re.search(r"Q(\d)", trimestre)
        if not quarter_match:
            raise HTTPException(status_code=400, detail=f"Cannot parse quarter from '{trimestre}'")
        quarter_num = int(quarter_match.group(1))  # 1-4

        # Legacy columns (0-indexed):
        # A=empresa, B-E=EF Q1-Q4, F=obs EF, G-J=IT Q1-Q4, K=obs IT
        ef_col = 1 + quarter_num  # B=Q1, C=Q2, D=Q3, E=Q4 → 2,3,4,5
        it_col = 6 + quarter_num  # G=Q1, H=Q2, I=Q3, J=Q4 → 7,8,9,10
        obs_ef_col = 6   # F
        obs_it_col = 11  # K

        # Data starts at row 4
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
            if not row or not row[0]:
                continue

            empresa_name = str(row[0]).strip()
            total_procesados += 1

            # Skip non-Madrid companies
            upper_name = empresa_name.upper()
            if any(upper_name.endswith(f" {suf}") or upper_name.endswith(f"_{suf}") for suf in LEGACY_SKIP_SUFFIXES):
                continue

            # Normalize (remove MAD suffix)
            normalized = normalize_empresa_name(empresa_name)

            # Find empresa
            match = empresas.get(normalized)
            if not match:
                # Fuzzy match
                for key, val in empresas.items():
                    if normalized in key or key in normalized:
                        match = val
                        break
            if not match:
                warnings.append(f"Fila {row_idx}: Empresa '{empresa_name}' no encontrada")
                continue

            # Parse EF and IT frequencies
            ef_freq = parse_legacy_frequency(row[ef_col - 1] if len(row) > ef_col - 1 else None)
            it_freq = parse_legacy_frequency(row[it_col - 1] if len(row) > it_col - 1 else None)
            total_freq = ef_freq + it_freq

            # Derive tipo
            if ef_freq > 0 and it_freq > 0:
                tipo = "AMBAS"
            elif ef_freq > 0:
                tipo = "EF"
            elif it_freq > 0:
                tipo = "IT"
            else:
                tipo = None

            # Merge observations
            obs_ef = str(row[obs_ef_col - 1]).strip() if len(row) > obs_ef_col - 1 and row[obs_ef_col - 1] else ""
            obs_it = str(row[obs_it_col - 1]).strip() if len(row) > obs_it_col - 1 and row[obs_it_col - 1] else ""
            notas_parts = []
            if obs_ef:
                notas_parts.append(f"EF: {obs_ef}")
            if obs_it:
                notas_parts.append(f"IT: {obs_it}")
            notas = " | ".join(notas_parts) if notas_parts else None

            preview.append(ImportPreviewItem(
                empresa_id=match["id"],
                nombre=match["nombre"],
                frecuencia=total_freq,
                tipo=tipo,
                notas=notas,
                detalle_ef=ef_freq if ef_freq > 0 else None,
                detalle_it=it_freq if it_freq > 0 else None,
            ))

    aplicados = 0
    if not dry_run and preview:
        # Apply changes
        for item in preview:
            updates = ['"frecuenciaSolicitada" = :freq', '"updatedAt" = NOW()']
            params = {"eid": item.empresa_id, "tri": trimestre, "freq": item.frecuencia}

            if item.tipo:
                updates.append('"tipoParticipacion" = :tipo')
                params["tipo"] = item.tipo

            if item.notas:
                updates.append("notas = :notas")
                params["notas"] = item.notas

            query = f"""
                UPDATE "configTrimestral"
                SET {', '.join(updates)}
                WHERE "empresaId" = :eid AND trimestre = :tri
            """
            result = await db.execute(text(query), params)
            if result.rowcount > 0:
                aplicados += 1

        await db.commit()

    return ImportarExcelResponse(
        trimestre=trimestre,
        formato_detectado=formato,
        total_procesados=total_procesados,
        aplicados=aplicados,
        preview=[p.model_dump() for p in preview],
        warnings=warnings,
        dry_run=dry_run,
    )


@router.put("/{trimestre}/batch")
async def actualizar_configs_batch(
    trimestre: str,
    body: ConfigBatchUpdateInput,
    db: AsyncSession = Depends(get_db),
):
    """
    Actualiza múltiples configuraciones en una sola llamada.
    Ideal para guardar cambios de la tabla editable.
    """
    updated = 0
    errors = []

    for item in body.updates:
        try:
            # Verificar que existe la config
            cfg_check = await db.execute(
                text("""
                    SELECT id FROM "configTrimestral"
                    WHERE "empresaId" = :eid AND trimestre = :tri
                """),
                {"eid": item.empresa_id, "tri": trimestre},
            )
            if not cfg_check.first():
                errors.append(f"Config para empresa {item.empresa_id} no encontrada")
                continue

            updates = []
            params = {"eid": item.empresa_id, "tri": trimestre}

            if item.tipo_participacion is not None:
                updates.append('"tipoParticipacion" = :tipo')
                params["tipo"] = item.tipo_participacion

            if item.escuela_propia is not None:
                updates.append('"escuelaPropia" = :escuela')
                params["escuela"] = item.escuela_propia

            if item.frecuencia_solicitada is not None:
                updates.append('"frecuenciaSolicitada" = :freq')
                params["freq"] = item.frecuencia_solicitada

            if item.disponibilidad_dias is not None:
                updates.append('"disponibilidadDias" = :dias')
                params["dias"] = item.disponibilidad_dias

            if item.turno_preferido is not None:
                updates.append('"turnoPreferido" = :turno')
                params["turno"] = item.turno_preferido if item.turno_preferido != "" else None

            if item.voluntarios_disponibles is not None:
                updates.append('"voluntariosDisponibles" = :vol')
                params["vol"] = item.voluntarios_disponibles

            if item.preferencias_taller is not None:
                updates.append('"preferenciasTaller" = :pref')
                params["pref"] = item.preferencias_taller

            if item.notas is not None:
                updates.append("notas = :notas")
                params["notas"] = item.notas

            if updates:
                updates.append('"updatedAt" = NOW()')
                query = f"""
                    UPDATE "configTrimestral"
                    SET {', '.join(updates)}
                    WHERE "empresaId" = :eid AND trimestre = :tri
                """
                await db.execute(text(query), params)
                updated += 1

        except Exception as e:
            errors.append(f"Empresa {item.empresa_id}: {str(e)}")

    await db.commit()
    return {"updated": updated, "errors": errors}


@router.post("/{trimestre}/inicializar")
async def inicializar_configs(
    trimestre: str,
    body: InicializarInput,
    db: AsyncSession = Depends(get_db),
):
    """
    Inicializa configuraciones para un trimestre nuevo.

    Si origen_trimestre se proporciona:
      - Clona las configs de ese trimestre (como el endpoint clonar-trimestre)

    Si no se proporciona:
      - Crea configs por defecto para todas las empresas activas
    """
    warnings: list[str] = []
    clonadas = 0
    nuevas = 0

    if body.origen_trimestre:
        # ── Modo clonar ──────────────────────────────────────
        rows = await db.execute(
            text("""
                SELECT
                    ct."empresaId",
                    e.nombre,
                    e.activa,
                    ct."tipoParticipacion",
                    ct."escuelaPropia",
                    ct."turnoPreferido",
                    ct."frecuenciaSolicitada",
                    ct."disponibilidadDias",
                    ct."voluntariosDisponibles",
                    ct."preferenciasTaller",
                    ct.notas
                FROM "configTrimestral" ct
                JOIN empresa e ON e.id = ct."empresaId"
                WHERE ct.trimestre = :origen
                ORDER BY e.nombre
            """),
            {"origen": body.origen_trimestre},
        )
        configs = [dict(r) for r in rows.mappings().all()]

        if not configs:
            raise HTTPException(
                status_code=404,
                detail=f"No hay configuraciones para el trimestre {body.origen_trimestre}",
            )

        saltadas = []
        for cfg in configs:
            if not cfg["activa"]:
                saltadas.append(cfg["nombre"])
                continue

            result = await db.execute(
                text("""
                    INSERT INTO "configTrimestral" (
                        "empresaId", trimestre, "tipoParticipacion",
                        "escuelaPropia", "disponibilidadDias",
                        "turnoPreferido", "frecuenciaSolicitada",
                        "voluntariosDisponibles", "preferenciasTaller",
                        notas, "updatedAt"
                    )
                    VALUES (
                        :eid, :destino, :tipo,
                        :escuela, :dias,
                        :turno, :freq,
                        :vol, :pref,
                        :notas, NOW()
                    )
                    ON CONFLICT ("empresaId", trimestre) DO NOTHING
                """),
                {
                    "eid": cfg["empresaId"],
                    "destino": trimestre,
                    "tipo": cfg["tipoParticipacion"],
                    "escuela": cfg["escuelaPropia"] or False,
                    "dias": cfg["disponibilidadDias"] or "L,M,X,J,V",
                    "turno": cfg["turnoPreferido"],
                    "freq": cfg["frecuenciaSolicitada"],
                    "vol": cfg["voluntariosDisponibles"] or 0,
                    "pref": cfg["preferenciasTaller"],
                    "notas": cfg["notas"],
                },
            )
            if result.rowcount > 0:
                clonadas += 1

        if saltadas:
            warnings.append(
                f"{len(saltadas)} empresa(s) inactiva(s) no clonadas: {', '.join(saltadas[:5])}"
                + ("..." if len(saltadas) > 5 else "")
            )

    else:
        # ── Modo crear por defecto ───────────────────────────
        # Obtener empresas activas que no tienen config para este trimestre
        rows = await db.execute(
            text("""
                SELECT e.id, e.nombre, e.tipo, e."turnoPreferido"
                FROM empresa e
                WHERE e.activa = true
                AND NOT EXISTS (
                    SELECT 1 FROM "configTrimestral" ct
                    WHERE ct."empresaId" = e.id AND ct.trimestre = :tri
                )
                ORDER BY e.nombre
            """),
            {"tri": trimestre},
        )
        empresas = [dict(r) for r in rows.mappings().all()]

        for emp in empresas:
            await db.execute(
                text("""
                    INSERT INTO "configTrimestral" (
                        "empresaId", trimestre, "tipoParticipacion",
                        "escuelaPropia", "disponibilidadDias",
                        "turnoPreferido", "voluntariosDisponibles",
                        "updatedAt"
                    )
                    VALUES (
                        :eid, :tri, :tipo,
                        false, 'L,M,X,J,V',
                        :turno, 0,
                        NOW()
                    )
                """),
                {
                    "eid": emp["id"],
                    "tri": trimestre,
                    "tipo": emp["tipo"] or "AMBAS",
                    "turno": emp["turnoPreferido"],
                },
            )
            nuevas += 1

    await db.commit()

    # Contar total de configs para el trimestre
    count_result = await db.execute(
        text('SELECT COUNT(*) FROM "configTrimestral" WHERE trimestre = :tri'),
        {"tri": trimestre},
    )
    total = count_result.scalar() or 0

    return {
        "trimestre": trimestre,
        "total_configs": total,
        "clonadas": clonadas,
        "nuevas": nuevas,
        "warnings": warnings,
    }


# NOTE: This route MUST come AFTER specific routes like /resumen, /batch, /inicializar
# because it catches all remaining paths with {empresa_id}
@router.put("/{trimestre}/{empresa_id}")
async def actualizar_config(
    trimestre: str,
    empresa_id: int,
    body: ConfigTrimestralUpdate,
    db: AsyncSession = Depends(get_db),
):
    """
    Actualiza la configuración trimestral de una empresa.
    Si no existe, la crea con valores por defecto.
    """
    # Verificar que la empresa existe
    emp_check = await db.execute(
        text("SELECT id, nombre FROM empresa WHERE id = :eid"),
        {"eid": empresa_id},
    )
    empresa = emp_check.mappings().first()
    if not empresa:
        raise HTTPException(status_code=404, detail=f"Empresa {empresa_id} no encontrada")

    # Verificar si existe la config
    cfg_check = await db.execute(
        text("""
            SELECT id FROM "configTrimestral"
            WHERE "empresaId" = :eid AND trimestre = :tri
        """),
        {"eid": empresa_id, "tri": trimestre},
    )
    existing = cfg_check.first()

    if existing:
        # Update existente
        updates = []
        params = {"eid": empresa_id, "tri": trimestre}

        if body.tipo_participacion is not None:
            updates.append('"tipoParticipacion" = :tipo')
            params["tipo"] = body.tipo_participacion

        if body.escuela_propia is not None:
            updates.append('"escuelaPropia" = :escuela')
            params["escuela"] = body.escuela_propia

        if body.frecuencia_solicitada is not None:
            updates.append('"frecuenciaSolicitada" = :freq')
            params["freq"] = body.frecuencia_solicitada

        if body.disponibilidad_dias is not None:
            updates.append('"disponibilidadDias" = :dias')
            params["dias"] = body.disponibilidad_dias

        if body.turno_preferido is not None:
            updates.append('"turnoPreferido" = :turno')
            params["turno"] = body.turno_preferido if body.turno_preferido != "" else None

        if body.voluntarios_disponibles is not None:
            updates.append('"voluntariosDisponibles" = :vol')
            params["vol"] = body.voluntarios_disponibles

        if body.preferencias_taller is not None:
            updates.append('"preferenciasTaller" = :pref')
            params["pref"] = body.preferencias_taller

        if body.notas is not None:
            updates.append("notas = :notas")
            params["notas"] = body.notas

        if updates:
            updates.append('"updatedAt" = NOW()')
            query = f"""
                UPDATE "configTrimestral"
                SET {', '.join(updates)}
                WHERE "empresaId" = :eid AND trimestre = :tri
            """
            await db.execute(text(query), params)
    else:
        # Crear nueva config
        await db.execute(
            text("""
                INSERT INTO "configTrimestral" (
                    "empresaId", trimestre, "tipoParticipacion",
                    "escuelaPropia", "frecuenciaSolicitada", "disponibilidadDias",
                    "turnoPreferido", "voluntariosDisponibles", "preferenciasTaller",
                    notas, "updatedAt"
                )
                VALUES (
                    :eid, :tri, :tipo, :escuela, :freq, :dias,
                    :turno, :vol, :pref, :notas, NOW()
                )
            """),
            {
                "eid": empresa_id,
                "tri": trimestre,
                "tipo": body.tipo_participacion or "AMBAS",
                "escuela": body.escuela_propia or False,
                "freq": body.frecuencia_solicitada,
                "dias": body.disponibilidad_dias or "L,M,X,J,V",
                "turno": body.turno_preferido if body.turno_preferido != "" else None,
                "vol": body.voluntarios_disponibles or 0,
                "pref": body.preferencias_taller,
                "notas": body.notas,
            },
        )

    await db.commit()

    # Devolver config actualizada
    result = await db.execute(
        text("""
            SELECT
                ct.id,
                ct."empresaId" AS empresa_id,
                e.nombre AS empresa_nombre,
                ct."tipoParticipacion" AS tipo_participacion,
                ct."escuelaPropia" AS escuela_propia,
                ct."frecuenciaSolicitada" AS frecuencia_solicitada,
                ct."disponibilidadDias" AS disponibilidad_dias,
                ct."turnoPreferido" AS turno_preferido,
                ct."voluntariosDisponibles" AS voluntarios_disponibles,
                ct."preferenciasTaller" AS preferencias_taller,
                ct.notas
            FROM "configTrimestral" ct
            JOIN empresa e ON e.id = ct."empresaId"
            WHERE ct."empresaId" = :eid AND ct.trimestre = :tri
        """),
        {"eid": empresa_id, "tri": trimestre},
    )
    config = result.mappings().first()

    return {"config": dict(config) if config else None}
